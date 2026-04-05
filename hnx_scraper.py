"""
HNX Bond Scraper — Desktop App
================================
Yêu cầu:
    pip install ttkbootstrap requests beautifulsoup4 lxml openpyxl tkcalendar

Chạy:
    python hnx_app.py
"""

import os
import sys
import threading
import calendar
from datetime import date, datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

# ── GUI ──────────────────────────────────────────────────────────────────────
import tkinter as tk
from tkinter import filedialog, messagebox
try:
    import ttkbootstrap as ttk
    from ttkbootstrap.constants import *
    HAS_BOOTSTRAP = True
except ImportError:
    import tkinter.ttk as ttk
    HAS_BOOTSTRAP = False

try:
    from tkcalendar import DateEntry
    HAS_CALENDAR = True
except ImportError:
    HAS_CALENDAR = False

# ── Core scraper (same logic as hnx_scraper.py) ──────────────────────────────
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

BASE_URL = "https://hnx.vn"
PAGE_REF = "https://hnx.vn/vi-vn/trai-phieu/ket-qua-gd-trong-ngay.html"

API_ENDPOINTS = {
    "outright":     "/ModuleReportBonds/Bond_KQGD_TrongNgay/GetTradingOutRightInDay",
    "repo_buy":     "/ModuleReportBonds/Bond_KQGD_TrongNgay/GetTradingReposInDay",
    "repo_sell":    "/ModuleReportBonds/Bond_KQGD_TrongNgay/GetTradingSaleAndRepurchaseInDay",
    "bond_lending": "/ModuleReportBonds/Bond_KQGD_TrongNgay/GetTradingLoanInDay",
}

TRANSACTION_TYPES = {
    "outright":     "Giao dịch Outright",
    "repo_buy":     "Giao dịch Mua Bán Lại",
    "repo_sell":    "Giao dịch Bán và Mua Lại",
    "bond_lending": "Giao dịch Vay Trái Phiếu",
}

HTTP_HEADERS = {
    "Accept":           "*/*",
    "Accept-Language":  "vi-VN,vi;q=0.9",
    "Content-Type":     "application/x-www-form-urlencoded",
    "Origin":           BASE_URL,
    "Referer":          PAGE_REF,
    "User-Agent":       "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "X-Requested-With": "XMLHttpRequest",
}

RECORDS_PER_PAGE = 100
_HEADER_CACHE: dict = {}

FALLBACK_HEADERS = {
    "outright":     ["STT","Ngày giao kết giao dịch","Ngày BĐGD","Kỳ hạn còn lại","Mã TP","Tiền tệ","Giá yết (đồng)","Lợi suất (%/năm)","KLGD","GTGD (đồng)","Ngày thanh toán"],
    "repo_buy":     ["STT","Ngày giao kết giao dịch","Ngày BĐGD","Kỳ hạn còn lại","Mã TP","Tiền tệ","Giá yết mua lại (đồng)","Lãi suất (%/năm)","KLGD","GTGD (đồng)","Ngày mua lại","Ngày thanh toán"],
    "repo_sell":    ["STT","Ngày giao kết giao dịch","Ngày BĐGD","Kỳ hạn còn lại","Mã TP","Tiền tệ","Giá yết bán lại (đồng)","Lãi suất (%/năm)","KLGD","GTGD (đồng)","Ngày bán lại","Ngày thanh toán"],
    "bond_lending": ["STT","Ngày giao kết giao dịch","Ngày BĐGD","Kỳ hạn còn lại","Mã TP","Tiền tệ","Giá yết (đồng)","Phí vay (%/năm)","KLGD","GTGD (đồng)","Ngày hoàn trả","Ngày thanh toán"],
}


def make_session():
    s = requests.Session()
    s.headers.update(HTTP_HEADERS)
    s.verify = False
    try:
        s.get(PAGE_REF, timeout=15)
    except Exception:
        pass
    return s


def try_numeric(val):
    if not isinstance(val, str):
        return val
    v = val.strip()
    if not v:
        return v
    dots, commas = v.count("."), v.count(",")
    if commas == 1 and dots == 0:
        try: return float(v.replace(",", "."))
        except ValueError: pass
    elif commas == 0 and dots >= 1:
        try: return int(v.replace(".", ""))
        except ValueError: pass
    elif commas == 1 and dots >= 1:
        try: return float(v.replace(".", "").replace(",", "."))
        except ValueError: pass
    elif commas == 0 and dots == 0:
        try: return int(v)
        except ValueError: pass
    return v


def parse_html_table(html: str) -> list[dict]:
    import re
    soup = BeautifulSoup(html, "lxml")
    table = soup.find("table")
    if not table:
        return []
    headers = [th.get_text(strip=True) for th in table.select("thead th, thead td")]
    rows = []
    for tr in table.select("tbody tr"):
        cells = [td.get_text(strip=True).replace("\xa0", "") for td in tr.find_all("td")]
        if not any(cells):
            continue
        if headers and len(cells) >= len(headers):
            rows.append(dict(zip(headers, cells[:len(headers)])))
        elif cells:
            rows.append({f"col_{i}": v for i, v in enumerate(cells)})
    return rows


def get_total_from_html(html: str) -> int:
    import re
    m = re.search(r"Tổng số\s+(\d+)\s+bản ghi", html)
    return int(m.group(1)) if m else 0


def fetch_all_pages(session, tx_type: str, trade_date: str) -> list[dict]:
    import math
    url  = BASE_URL + API_ENDPOINTS[tx_type]
    data = {"p_keysearch": trade_date + "|", "pColOrder": "col_c",
            "pOrderType": "ASC", "pCurrentPage": "1",
            "pRecordOnPage": str(RECORDS_PER_PAGE), "pIsSearch": "1", "pIsChangeTab": "0"}
    try:
        resp  = session.post(url, data=data, timeout=20)
        resp.raise_for_status()
        html  = resp.text
        total = get_total_from_html(html)
        rows  = parse_html_table(html)
    except Exception as e:
        return []

    if total > RECORDS_PER_PAGE:
        for p in range(2, math.ceil(total / RECORDS_PER_PAGE) + 1):
            try:
                data["pCurrentPage"] = str(p)
                r2 = session.post(url, data=data, timeout=20)
                rows.extend(parse_html_table(r2.text))
            except Exception:
                break

    date_key = next((k for k in (rows[0].keys() if rows else [])
                     if any(x in k.lower() for x in ["bđgd","bdgd","ngày b"])), None)
    if date_key:
        rows = [r for r in rows if str(r.get(date_key,"")).strip() == trade_date]
    return rows


def fetch_day(session, trade_date: str, types_to_fetch: list) -> dict:
    return {tx: fetch_all_pages(session, tx, trade_date) for tx in types_to_fetch}


def date_range(start: date, end: date) -> list[date]:
    days, cur = [], start
    while cur <= end:
        if cur.weekday() < 5:
            days.append(cur)
        cur += timedelta(days=1)
    return days


def dates_for_month(year: int, month: int) -> list[date]:
    _, last = calendar.monthrange(year, month)
    return date_range(date(year, month, 1), date(year, month, last))


# ── Excel helpers ─────────────────────────────────────────────────────────────
NAVY, BLUE, LT_BLUE = "1F4E79", "2E75B6", "EBF3FB"

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(cell, text, bg=BLUE, fc="FFFFFF", bold=True, size=10):
    cell.value = text
    cell.font  = Font(name="Arial", bold=bold, size=size, color=fc)
    cell.fill  = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border()

def get_or_create_sheet(wb, sheet_name, tx_type, keys):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        return ws, ws.max_row + 1
    ws = wb.create_sheet(sheet_name)
    ws.row_dimensions[1].height = 28
    for ci, k in enumerate(keys, 1):
        _hdr(ws.cell(1, ci), k)
        ws.column_dimensions[get_column_letter(ci)].width = max(14, len(k)+3)
    ws.freeze_panes = "A2"
    return ws, 2

def append_records(wb, sheet_name, records, tx_type):
    if records:
        keys = list(records[0].keys())
        _HEADER_CACHE[tx_type] = keys
    else:
        keys = _HEADER_CACHE.get(tx_type) or FALLBACK_HEADERS.get(tx_type) or FALLBACK_HEADERS["outright"]
    ws, start_row = get_or_create_sheet(wb, sheet_name, tx_type, keys)
    for ri, rec in enumerate(records):
        r    = start_row + ri
        fill = PatternFill("solid", start_color=LT_BLUE if r % 2 == 0 else "FFFFFF")
        for ci, k in enumerate(keys, 1):
            cell = ws.cell(r, ci)
            val  = try_numeric(rec.get(k, ""))
            cell.value, cell.font, cell.fill, cell.border = val, Font(name="Arial", size=10), fill, _border()
            if isinstance(val, (int, float)):
                cell.number_format = "#,##0.####"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

def export_excel(all_fetched, output_path, date_range_label, types_to_fetch):
    wb = Workbook()
    wb.remove(wb.active)
    summary = {tx: {"days": 0, "records": 0} for tx in types_to_fetch}
    for trade_date, fetched in all_fetched:
        for tx_type in types_to_fetch:
            records = fetched.get(tx_type, [])
            if records:
                append_records(wb, TRANSACTION_TYPES[tx_type][:28], records, tx_type)
                summary[tx_type]["days"]    += 1
                summary[tx_type]["records"] += len(records)
    # Ensure empty sheets
    for tx_type in types_to_fetch:
        sname = TRANSACTION_TYPES[tx_type][:28]
        if sname not in wb.sheetnames:
            keys = _HEADER_CACHE.get(tx_type) or FALLBACK_HEADERS.get(tx_type) or FALLBACK_HEADERS["outright"]
            ws = wb.create_sheet(sname)
            ws.row_dimensions[1].height = 28
            for ci, k in enumerate(keys, 1):
                _hdr(ws.cell(1, ci), k)
                ws.column_dimensions[get_column_letter(ci)].width = max(14, len(k)+3)
            ws.freeze_panes = "A2"
    # Cover sheet
    ws0 = wb.create_sheet("Tổng quan", 0)
    for col, w in zip("ABCD", [30,28,18,14]): ws0.column_dimensions[col].width = w
    ws0.merge_cells("A1:D1"); _hdr(ws0["A1"], "KẾT QUẢ GIAO DỊCH TRÁI PHIẾU – HNX", bg=NAVY, size=13)
    ws0.row_dimensions[1].height = 30
    ws0.merge_cells("A2:D2")
    _hdr(ws0["A2"], f"Kỳ: {date_range_label}  |  Tạo lúc: {datetime.now():%d/%m/%Y %H:%M}  |  Nguồn: hnx.vn", bg=BLUE, size=9, bold=False)
    ws0.row_dimensions[2].height = 16
    ws0.append([])
    for ci, h in enumerate(["Loại giao dịch","Sheet","Số ngày có GD","Tổng bản ghi"], 1):
        _hdr(ws0.cell(4, ci), h)
    for i, tx in enumerate(types_to_fetch):
        r    = 5 + i
        s    = summary[tx]
        fill = PatternFill("solid", start_color=LT_BLUE if i%2==0 else "FFFFFF")
        for ci, (val, align) in enumerate([(TRANSACTION_TYPES[tx],"left"),(TRANSACTION_TYPES[tx][:28],"center"),(s["days"],"right"),(s["records"],"right")],1):
            cell = ws0.cell(r, ci); cell.value=val; cell.font=Font(name="Arial",size=10); cell.fill=fill; cell.border=_border()
            cell.alignment=Alignment(horizontal=align,vertical="center")
            if ci in (3,4): cell.number_format="#,##0"
    wb.save(output_path)


# ════════════════════════════════════════════════════════════════════════════════
#  GUI Application
# ════════════════════════════════════════════════════════════════════════════════

class HNXApp:
    def __init__(self, root):
        self.root = root
        self.root.title("HNX Bond Scraper")
        self.root.resizable(False, False)
        self._running = False

        if HAS_BOOTSTRAP:
            style = ttk.Style(theme="cosmo")
            self.root.configure(bg=style.colors.bg)

        self._build_ui()

    # ── UI Builder ────────────────────────────────────────────────────────────
    def _build_ui(self):
        pad = dict(padx=12, pady=6)

        # ── Title ──
        title = ttk.Label(self.root, text="HNX Bond Scraper",
                          font=("Arial", 16, "bold"))
        title.grid(row=0, column=0, columnspan=2, pady=(16, 4))

        sub = ttk.Label(self.root, text="Kết quả giao dịch trái phiếu  •  hnx.vn",
                        font=("Arial", 9), foreground="gray")
        sub.grid(row=1, column=0, columnspan=2, pady=(0, 12))

        # ── Date mode ────────────────────────────────────────────────────────
        lf_mode = ttk.LabelFrame(self.root, text="  Chế độ ngày  ", padding=10)
        lf_mode.grid(row=2, column=0, columnspan=2, sticky="ew", **pad)

        self.date_mode = tk.StringVar(value="single")
        modes = [("Một ngày", "single"), ("Từ ngày – đến ngày", "range"), ("Cả tháng", "month")]
        for i, (lbl, val) in enumerate(modes):
            rb = ttk.Radiobutton(lf_mode, text=lbl, variable=self.date_mode,
                                 value=val, command=self._on_mode_change)
            rb.grid(row=0, column=i, padx=16, pady=4, sticky="w")

        # ── Date inputs ──────────────────────────────────────────────────────
        lf_date = ttk.LabelFrame(self.root, text="  Ngày  ", padding=10)
        lf_date.grid(row=3, column=0, columnspan=2, sticky="ew", **pad)
        lf_date.columnconfigure(1, weight=1)
        lf_date.columnconfigure(3, weight=1)

        # Single date
        self.lbl_single = ttk.Label(lf_date, text="Ngày BĐGD:")
        self.lbl_single.grid(row=0, column=0, sticky="w", padx=(0,8))
        self.date_single = self._make_date_entry(lf_date, date.today())
        self.date_single.grid(row=0, column=1, sticky="w")

        # Range
        self.lbl_from = ttk.Label(lf_date, text="Từ ngày:")
        self.lbl_from.grid(row=1, column=0, sticky="w", padx=(0,8), pady=(8,0))
        self.date_from = self._make_date_entry(lf_date, date.today().replace(day=1))
        self.date_from.grid(row=1, column=1, sticky="w", pady=(8,0))

        self.lbl_to = ttk.Label(lf_date, text="Đến ngày:")
        self.lbl_to.grid(row=1, column=2, sticky="w", padx=(16,8), pady=(8,0))
        self.date_to = self._make_date_entry(lf_date, date.today())
        self.date_to.grid(row=1, column=3, sticky="w", pady=(8,0))

        # Month picker
        self.lbl_month = ttk.Label(lf_date, text="Tháng/Năm:")
        self.lbl_month.grid(row=2, column=0, sticky="w", padx=(0,8), pady=(8,0))

        month_frame = ttk.Frame(lf_date)
        month_frame.grid(row=2, column=1, sticky="w", pady=(8,0))

        months = [f"{i:02d}" for i in range(1, 13)]
        self.month_var = tk.StringVar(value=f"{date.today().month:02d}")
        self.month_cb = ttk.Combobox(month_frame, textvariable=self.month_var,
                                     values=months, width=4, state="readonly")
        self.month_cb.pack(side="left")
        ttk.Label(month_frame, text="/").pack(side="left", padx=4)
        self.year_var = tk.StringVar(value=str(date.today().year))
        self.year_sb = ttk.Spinbox(month_frame, from_=2020, to=2035,
                                   textvariable=self.year_var, width=6)
        self.year_sb.pack(side="left")

        self._on_mode_change()  # set initial visibility

        # ── Transaction types ─────────────────────────────────────────────────
        lf_tx = ttk.LabelFrame(self.root, text="  Loại giao dịch  ", padding=10)
        lf_tx.grid(row=4, column=0, columnspan=2, sticky="ew", **pad)

        self.tx_vars = {}
        tx_items = list(TRANSACTION_TYPES.items())
        for i, (key, label) in enumerate(tx_items):
            var = tk.BooleanVar(value=True)
            self.tx_vars[key] = var
            cb = ttk.Checkbutton(lf_tx, text=label, variable=var)
            cb.grid(row=i//2, column=i%2, sticky="w", padx=16, pady=3)

        # Select all / none buttons
        btn_frame = ttk.Frame(lf_tx)
        btn_frame.grid(row=2, column=0, columnspan=2, pady=(8,0))
        ttk.Button(btn_frame, text="Chọn tất cả",  width=14,
                   command=lambda: [v.set(True)  for v in self.tx_vars.values()]).pack(side="left", padx=4)
        ttk.Button(btn_frame, text="Bỏ chọn tất cả", width=14,
                   command=lambda: [v.set(False) for v in self.tx_vars.values()]).pack(side="left", padx=4)

        # ── Output file ──────────────────────────────────────────────────────
        lf_out = ttk.LabelFrame(self.root, text="  File xuất  ", padding=10)
        lf_out.grid(row=5, column=0, columnspan=2, sticky="ew", **pad)
        lf_out.columnconfigure(0, weight=1)

        out_row = ttk.Frame(lf_out)
        out_row.grid(row=0, column=0, sticky="ew")
        out_row.columnconfigure(0, weight=1)

        self.output_var = tk.StringVar(value=os.path.join(
            os.path.expanduser("~"), "Desktop",
            f"HNX_TraiPhieu_{date.today():%d%m%Y}.xlsx"
        ))
        ttk.Entry(out_row, textvariable=self.output_var).grid(row=0, column=0, sticky="ew", padx=(0,8))
        ttk.Button(out_row, text="📁 Chọn...", width=10,
                   command=self._browse_output).grid(row=0, column=1)

        # ── Workers ──────────────────────────────────────────────────────────
        workers_frame = ttk.Frame(self.root)
        workers_frame.grid(row=6, column=0, columnspan=2, sticky="w", padx=12, pady=(4,0))
        ttk.Label(workers_frame, text="Số ngày song song:").pack(side="left")
        self.workers_var = tk.IntVar(value=4)
        ttk.Spinbox(workers_frame, from_=1, to=8, textvariable=self.workers_var,
                    width=4).pack(side="left", padx=8)
        ttk.Label(workers_frame, text="(tăng để nhanh hơn, giảm nếu bị chặn IP)",
                  foreground="gray", font=("Arial",8)).pack(side="left")

        # ── Run button ───────────────────────────────────────────────────────
        self.btn_run = ttk.Button(self.root, text="▶  Bắt đầu tải dữ liệu",
                                  command=self._start, width=30,
                                  bootstyle="success" if HAS_BOOTSTRAP else "")
        self.btn_run.grid(row=7, column=0, columnspan=2, pady=(12, 6))

        # ── Progress ─────────────────────────────────────────────────────────
        self.progress = ttk.Progressbar(self.root, length=440, mode="determinate")
        self.progress.grid(row=8, column=0, columnspan=2, padx=12, pady=(0,4))

        self.status_var = tk.StringVar(value="Sẵn sàng")
        ttk.Label(self.root, textvariable=self.status_var,
                  foreground="gray").grid(row=9, column=0, columnspan=2, pady=(0,4))

        # ── Log ──────────────────────────────────────────────────────────────
        lf_log = ttk.LabelFrame(self.root, text="  Log  ", padding=6)
        lf_log.grid(row=10, column=0, columnspan=2, sticky="nsew", padx=12, pady=(0,12))
        self.root.rowconfigure(10, weight=1)
        self.root.columnconfigure(0, weight=1)

        self.log_text = tk.Text(lf_log, height=10, width=58, state="disabled",
                                font=("Consolas", 9), bg="#1e1e1e", fg="#d4d4d4",
                                insertbackground="white", relief="flat")
        scroll = ttk.Scrollbar(lf_log, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scroll.set)
        self.log_text.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")

        self.root.geometry("500x680")

    # ── Date entry factory ────────────────────────────────────────────────────
    def _make_date_entry(self, parent, initial: date):
        if HAS_CALENDAR:
            e = DateEntry(parent, date_pattern="dd/MM/yyyy",
                          year=initial.year, month=initial.month, day=initial.day,
                          width=12, background="darkblue", foreground="white",
                          borderwidth=2)
            return e
        else:
            var = tk.StringVar(value=initial.strftime("%d/%m/%Y"))
            e   = ttk.Entry(parent, textvariable=var, width=13)
            e.get  = var.get   # duck-type to match DateEntry
            e.set  = var.set
            return e

    def _get_date(self, widget) -> date:
        """Get date from either DateEntry or plain Entry."""
        if HAS_CALENDAR and isinstance(widget, DateEntry):
            return widget.get_date()
        raw = widget.get().strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try: return datetime.strptime(raw, fmt).date()
            except ValueError: pass
        raise ValueError(f"Ngày không hợp lệ: {raw!r}")

    # ── Mode change ───────────────────────────────────────────────────────────
    def _on_mode_change(self):
        mode = self.date_mode.get()
        single_widgets = [self.lbl_single, self.date_single]
        range_widgets  = [self.lbl_from, self.date_from, self.lbl_to, self.date_to]
        month_widgets  = [self.lbl_month, self.month_cb, self.year_sb]

        for w in single_widgets: w.grid_remove() if mode != "single" else w.grid()
        for w in range_widgets:  w.grid_remove() if mode != "range"  else w.grid()
        for w in [self.lbl_month]:
            w.grid_remove() if mode != "month" else w.grid()

        # month combobox and spinbox are inside a frame — show/hide parent frame
        try:
            fr = self.month_cb.master
            fr.grid_remove() if mode != "month" else fr.grid()
        except Exception:
            pass

    # ── Browse output ─────────────────────────────────────────────────────────
    def _browse_output(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=os.path.basename(self.output_var.get()),
            initialdir=os.path.dirname(self.output_var.get()),
        )
        if path:
            self.output_var.set(path)

    # ── Logging ───────────────────────────────────────────────────────────────
    def _log(self, msg: str, tag: str = ""):
        def _do():
            self.log_text.configure(state="normal")
            prefix = {"ok": "✅ ", "err": "❌ ", "warn": "⚠️  ", "info": "ℹ️  "}.get(tag, "")
            self.log_text.insert("end", prefix + msg + "\n")
            self.log_text.see("end")
            self.log_text.configure(state="disabled")
        self.root.after(0, _do)

    def _set_status(self, msg: str):
        self.root.after(0, lambda: self.status_var.set(msg))

    def _set_progress(self, val: float):
        self.root.after(0, lambda: self.progress.configure(value=val))

    # ── Start ─────────────────────────────────────────────────────────────────
    def _start(self):
        if self._running:
            return

        # Validate types
        types_to_fetch = [k for k, v in self.tx_vars.items() if v.get()]
        if not types_to_fetch:
            messagebox.showwarning("Thiếu thông tin", "Vui lòng chọn ít nhất một loại giao dịch.")
            return

        # Validate output
        output = self.output_var.get().strip()
        if not output:
            messagebox.showwarning("Thiếu thông tin", "Vui lòng chọn file xuất.")
            return

        # Resolve dates
        try:
            mode = self.date_mode.get()
            if mode == "single":
                d      = self._get_date(self.date_single)
                dates  = [d] if d.weekday() < 5 else date_range(d, d)
                label  = d.strftime("%d/%m/%Y")
            elif mode == "range":
                d_from = self._get_date(self.date_from)
                d_to   = self._get_date(self.date_to)
                if d_to < d_from:
                    messagebox.showerror("Lỗi ngày", "Ngày đến phải lớn hơn ngày từ.")
                    return
                dates = date_range(d_from, d_to)
                label = f"{d_from:%d/%m/%Y} – {d_to:%d/%m/%Y}"
            else:
                m     = int(self.month_var.get())
                y     = int(self.year_var.get())
                dates = dates_for_month(y, m)
                label = f"Tháng {m:02d}/{y}"
        except ValueError as e:
            messagebox.showerror("Lỗi ngày", str(e))
            return

        if not dates:
            messagebox.showwarning("Không có ngày", "Khoảng thời gian không có ngày giao dịch (thứ 2–6).")
            return

        # Clear log
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")

        # Start worker thread
        self._running = True
        self.btn_run.configure(state="disabled",
                               text="⏳  Đang tải..." if not HAS_BOOTSTRAP else "⏳  Đang tải...")
        self.progress.configure(value=0, maximum=len(dates))

        threading.Thread(
            target=self._run_worker,
            args=(dates, types_to_fetch, output, label),
            daemon=True
        ).start()

    # ── Worker ────────────────────────────────────────────────────────────────
    def _run_worker(self, dates, types_to_fetch, output, label):
        try:
            self._log(f"Kỳ: {label}", "info")
            self._log(f"Số ngày: {len(dates)}  |  Loại GD: {len(types_to_fetch)}", "info")
            self._log(f"Output: {output}", "info")
            self._log("─" * 50)

            start_time  = __import__("time").time()
            all_fetched = []
            grand_total = 0
            skipped     = 0
            workers     = self.workers_var.get()

            def fetch_one(d: date):
                sess       = make_session()
                trade_date = d.strftime("%d/%m/%Y")
                fetched    = fetch_day(sess, trade_date, types_to_fetch)
                total      = sum(len(v) for v in fetched.values())
                return trade_date, fetched, total

            completed = 0
            results_map = {}

            with ThreadPoolExecutor(max_workers=workers) as ex:
                future_map = {ex.submit(fetch_one, d): d for d in dates}
                for future in as_completed(future_map):
                    d = future_map[future]
                    completed += 1
                    self._set_progress(completed)
                    try:
                        trade_date, fetched, total = future.result()
                        results_map[d] = (trade_date, fetched, total)
                        status = f"{total} bản ghi" if total else "không có dữ liệu"
                        self._log(f"[{completed}/{len(dates)}] {trade_date} → {status}")
                    except Exception as e:
                        results_map[d] = None
                        self._log(f"[{completed}/{len(dates)}] {d:%d/%m/%Y} → LỖI: {e}", "err")

            for d in dates:
                entry = results_map.get(d)
                if entry is None:
                    skipped += 1
                else:
                    trade_date, fetched, total = entry
                    if total == 0:
                        skipped += 1
                    else:
                        all_fetched.append((trade_date, fetched))
                        grand_total += total

            if not all_fetched:
                self._log("Không tìm thấy dữ liệu nào.", "warn")
                self._set_status("Không có dữ liệu")
                return

            self._log("─" * 50)
            self._log(f"Đang xuất Excel: {output}", "info")
            self._set_status("Đang xuất Excel...")

            export_excel(all_fetched, output, label, types_to_fetch)

            elapsed = __import__("time").time() - start_time
            self._log(f"Hoàn thành! {grand_total:,} bản ghi  |  {elapsed:.1f}s", "ok")
            if skipped:
                self._log(f"Bỏ qua {skipped} ngày không có dữ liệu", "warn")
            self._set_status(f"✅  Xong — {grand_total:,} bản ghi  ({elapsed:.1f}s)")

            # Ask to open file
            self.root.after(0, lambda: self._ask_open(output))

        except Exception as e:
            self._log(f"Lỗi: {e}", "err")
            self._set_status(f"❌  Lỗi: {e}")
        finally:
            self._running = False
            self.root.after(0, lambda: self.btn_run.configure(
                state="normal", text="▶  Bắt đầu tải dữ liệu"))

    def _ask_open(self, path):
        if messagebox.askyesno("Hoàn thành", f"Xuất thành công!\n\nMở file Excel ngay bây giờ?"):
            os.startfile(path) if sys.platform == "win32" else os.system(f'open "{path}"')


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if HAS_BOOTSTRAP:
        root = ttk.Window(themename="cosmo")
    else:
        root = tk.Tk()

    app = HNXApp(root)
    root.mainloop()