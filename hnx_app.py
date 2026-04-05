"""
HNX Bond Scraper — Desktop App
================================
Yêu cầu:
    pip install ttkbootstrap requests beautifulsoup4 lxml openpyxl tkcalendar

Chạy:
    python hnx_app.py
"""

import os
import sys
import threading
import calendar
from datetime import date, datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

# ── GUI ──────────────────────────────────────────────────────────────────────
import tkinter as tk
from tkinter import filedialog, messagebox
try:
    import ttkbootstrap as ttk
    from ttkbootstrap.constants import *
    HAS_BOOTSTRAP = True
except ImportError:
    import tkinter.ttk as ttk
    HAS_BOOTSTRAP = False

HAS_CALENDAR = False  # use built-in date spinbox instead

# ── Core scraper (same logic as hnx_scraper.py) ──────────────────────────────
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

BASE_URL = "https://hnx.vn"
PAGE_REF = "https://hnx.vn/vi-vn/trai-phieu/ket-qua-gd-trong-ngay.html"

API_ENDPOINTS = {
    "outright":     "/ModuleReportBonds/Bond_KQGD_TrongNgay/GetTradingOutRightInDay",
    "repo_buy":     "/ModuleReportBonds/Bond_KQGD_TrongNgay/GetTradingReposInDay",
    "repo_sell":    "/ModuleReportBonds/Bond_KQGD_TrongNgay/GetTradingSaleAndRepurchaseInDay",
    "bond_lending": "/ModuleReportBonds/Bond_KQGD_TrongNgay/GetTradingLoanInDay",
}

TRANSACTION_TYPES = {
    "outright":     "Giao dịch Outright",
    "repo_buy":     "Giao dịch Mua Bán Lại",
    "repo_sell":    "Giao dịch Bán và Mua Lại",
    "bond_lending": "Giao dịch Vay Trái Phiếu",
}

HTTP_HEADERS = {
    "Accept":           "*/*",
    "Accept-Language":  "vi-VN,vi;q=0.9",
    "Content-Type":     "application/x-www-form-urlencoded",
    "Origin":           BASE_URL,
    "Referer":          PAGE_REF,
    "User-Agent":       "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "X-Requested-With": "XMLHttpRequest",
}

RECORDS_PER_PAGE = 100
_HEADER_CACHE: dict = {}

FALLBACK_HEADERS = {
    "outright":     ["STT","Ngày giao kết giao dịch","Ngày BĐGD","Kỳ hạn còn lại","Mã TP","Tiền tệ","Giá yết (đồng)","Lợi suất (%/năm)","KLGD","GTGD (đồng)","Ngày thanh toán"],
    "repo_buy":     ["STT","Ngày giao kết giao dịch","Ngày BĐGD","Kỳ hạn còn lại","Mã TP","Tiền tệ","Giá yết mua lại (đồng)","Lãi suất (%/năm)","KLGD","GTGD (đồng)","Ngày mua lại","Ngày thanh toán"],
    "repo_sell":    ["STT","Ngày giao kết giao dịch","Ngày BĐGD","Kỳ hạn còn lại","Mã TP","Tiền tệ","Giá yết bán lại (đồng)","Lãi suất (%/năm)","KLGD","GTGD (đồng)","Ngày bán lại","Ngày thanh toán"],
    "bond_lending": ["STT","Ngày giao kết giao dịch","Ngày BĐGD","Kỳ hạn còn lại","Mã TP","Tiền tệ","Giá yết (đồng)","Phí vay (%/năm)","KLGD","GTGD (đồng)","Ngày hoàn trả","Ngày thanh toán"],
}


def make_session():
    s = requests.Session()
    s.headers.update(HTTP_HEADERS)
    s.verify = False
    try:
        s.get(PAGE_REF, timeout=15)
    except Exception:
        pass
    return s


def try_numeric(val):
    if not isinstance(val, str):
        return val
    v = val.strip()
    if not v:
        return v
    dots, commas = v.count("."), v.count(",")
    if commas == 1 and dots == 0:
        try: return float(v.replace(",", "."))
        except ValueError: pass
    elif commas == 0 and dots >= 1:
        try: return int(v.replace(".", ""))
        except ValueError: pass
    elif commas == 1 and dots >= 1:
        try: return float(v.replace(".", "").replace(",", "."))
        except ValueError: pass
    elif commas == 0 and dots == 0:
        try: return int(v)
        except ValueError: pass
    return v


def parse_html_table(html: str) -> list[dict]:
    import re
    soup = BeautifulSoup(html, "lxml")
    table = soup.find("table")
    if not table:
        return []
    headers = [th.get_text(strip=True) for th in table.select("thead th, thead td")]
    rows = []
    for tr in table.select("tbody tr"):
        cells = [td.get_text(strip=True).replace("\xa0", "") for td in tr.find_all("td")]
        if not any(cells):
            continue
        if headers and len(cells) >= len(headers):
            rows.append(dict(zip(headers, cells[:len(headers)])))
        elif cells:
            rows.append({f"col_{i}": v for i, v in enumerate(cells)})
    return rows


def get_total_from_html(html: str) -> int:
    import re
    m = re.search(r"Tổng số\s+(\d+)\s+bản ghi", html)
    return int(m.group(1)) if m else 0


def fetch_all_pages(session, tx_type: str, trade_date: str) -> list[dict]:
    import math
    url  = BASE_URL + API_ENDPOINTS[tx_type]
    data = {"p_keysearch": trade_date + "|", "pColOrder": "col_c",
            "pOrderType": "ASC", "pCurrentPage": "1",
            "pRecordOnPage": str(RECORDS_PER_PAGE), "pIsSearch": "1", "pIsChangeTab": "0"}
    try:
        resp  = session.post(url, data=data, timeout=20)
        resp.raise_for_status()
        html  = resp.text
        total = get_total_from_html(html)
        rows  = parse_html_table(html)
    except Exception as e:
        return []

    if total > RECORDS_PER_PAGE:
        for p in range(2, math.ceil(total / RECORDS_PER_PAGE) + 1):
            try:
                data["pCurrentPage"] = str(p)
                r2 = session.post(url, data=data, timeout=20)
                rows.extend(parse_html_table(r2.text))
            except Exception:
                break

    date_key = next((k for k in (rows[0].keys() if rows else [])
                     if any(x in k.lower() for x in ["bđgd","bdgd","ngày b"])), None)
    if date_key:
        rows = [r for r in rows if str(r.get(date_key,"")).strip() == trade_date]
    return rows


def fetch_day(session, trade_date: str, types_to_fetch: list) -> dict:
    return {tx: fetch_all_pages(session, tx, trade_date) for tx in types_to_fetch}


def date_range(start: date, end: date) -> list[date]:
    days, cur = [], start
    while cur <= end:
        if cur.weekday() < 5:
            days.append(cur)
        cur += timedelta(days=1)
    return days


def dates_for_month(year: int, month: int) -> list[date]:
    _, last = calendar.monthrange(year, month)
    return date_range(date(year, month, 1), date(year, month, last))


# ── Excel helpers ─────────────────────────────────────────────────────────────
NAVY, BLUE, LT_BLUE = "1F4E79", "2E75B6", "EBF3FB"

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(cell, text, bg=BLUE, fc="FFFFFF", bold=True, size=10):
    cell.value = text
    cell.font  = Font(name="Arial", bold=bold, size=size, color=fc)
    cell.fill  = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border()

def get_or_create_sheet(wb, sheet_name, tx_type, keys):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        return ws, ws.max_row + 1
    ws = wb.create_sheet(sheet_name)
    ws.row_dimensions[1].height = 28
    for ci, k in enumerate(keys, 1):
        _hdr(ws.cell(1, ci), k)
        ws.column_dimensions[get_column_letter(ci)].width = max(14, len(k)+3)
    ws.freeze_panes = "A2"
    return ws, 2

def append_records(wb, sheet_name, records, tx_type):
    if records:
        keys = list(records[0].keys())
        _HEADER_CACHE[tx_type] = keys
    else:
        keys = _HEADER_CACHE.get(tx_type) or FALLBACK_HEADERS.get(tx_type) or FALLBACK_HEADERS["outright"]
    ws, start_row = get_or_create_sheet(wb, sheet_name, tx_type, keys)
    for ri, rec in enumerate(records):
        r    = start_row + ri
        fill = PatternFill("solid", start_color=LT_BLUE if r % 2 == 0 else "FFFFFF")
        for ci, k in enumerate(keys, 1):
            cell = ws.cell(r, ci)
            val  = try_numeric(rec.get(k, ""))
            cell.value, cell.font, cell.fill, cell.border = val, Font(name="Arial", size=10), fill, _border()
            if isinstance(val, (int, float)):
                cell.number_format = "#,##0.####"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

def export_excel(all_fetched, output_path, date_range_label, types_to_fetch):
    wb = Workbook()
    wb.remove(wb.active)
    summary = {tx: {"days": 0, "records": 0} for tx in types_to_fetch}
    for trade_date, fetched in all_fetched:
        for tx_type in types_to_fetch:
            records = fetched.get(tx_type, [])
            if records:
                append_records(wb, TRANSACTION_TYPES[tx_type][:28], records, tx_type)
                summary[tx_type]["days"]    += 1
                summary[tx_type]["records"] += len(records)
    # Ensure empty sheets
    for tx_type in types_to_fetch:
        sname = TRANSACTION_TYPES[tx_type][:28]
        if sname not in wb.sheetnames:
            keys = _HEADER_CACHE.get(tx_type) or FALLBACK_HEADERS.get(tx_type) or FALLBACK_HEADERS["outright"]
            ws = wb.create_sheet(sname)
            ws.row_dimensions[1].height = 28
            for ci, k in enumerate(keys, 1):
                _hdr(ws.cell(1, ci), k)
                ws.column_dimensions[get_column_letter(ci)].width = max(14, len(k)+3)
            ws.freeze_panes = "A2"
    # Cover sheet
    ws0 = wb.create_sheet("Tổng quan", 0)
    for col, w in zip("ABCD", [30,28,18,14]): ws0.column_dimensions[col].width = w
    ws0.merge_cells("A1:D1"); _hdr(ws0["A1"], "KẾT QUẢ GIAO DỊCH TRÁI PHIẾU – HNX", bg=NAVY, size=13)
    ws0.row_dimensions[1].height = 30
    ws0.merge_cells("A2:D2")
    _hdr(ws0["A2"], f"Kỳ: {date_range_label}  |  Tạo lúc: {datetime.now():%d/%m/%Y %H:%M}  |  Nguồn: hnx.vn", bg=BLUE, size=9, bold=False)
    ws0.row_dimensions[2].height = 16
    ws0.append([])
    for ci, h in enumerate(["Loại giao dịch","Sheet","Số ngày có GD","Tổng bản ghi"], 1):
        _hdr(ws0.cell(4, ci), h)
    for i, tx in enumerate(types_to_fetch):
        r    = 5 + i
        s    = summary[tx]
        fill = PatternFill("solid", start_color=LT_BLUE if i%2==0 else "FFFFFF")
        for ci, (val, align) in enumerate([(TRANSACTION_TYPES[tx],"left"),(TRANSACTION_TYPES[tx][:28],"center"),(s["days"],"right"),(s["records"],"right")],1):
            cell = ws0.cell(r, ci); cell.value=val; cell.font=Font(name="Arial",size=10); cell.fill=fill; cell.border=_border()
            cell.alignment=Alignment(horizontal=align,vertical="center")
            if ci in (3,4): cell.number_format="#,##0"
    wb.save(output_path)





# ════════════════════════════════════════════════════════════════════════════════
#  Localisation
# ════════════════════════════════════════════════════════════════════════════════

LANG = {
"vi": {
    "app_title":"HNX Bond Scraper",
    "mode_single":"Một ngày","mode_range":"Từ ngày – đến ngày","mode_month":"Cả tháng",
    "lbl_date":"Ngày BĐGD:","lbl_from":"Từ:","lbl_to":"Đến:",
    "lbl_month":"Tháng:","lbl_year":"Năm:",
    "tx_title":"Loại giao dịch",
    "sel_all":"Chọn tất cả","clr_all":"Bỏ chọn",
    "out_lbl":"File Excel:","browse":"Chọn...",
    "workers_lbl":"Song song:","workers_hint":"ngày/lần",
    "run_btn":"▶  Bắt đầu tải","running_btn":"⏳  Đang tải...",
    "ready":"Sẵn sàng",
    "warn_type":"Chọn ít nhất một loại giao dịch.",
    "warn_out":"Chưa chọn file xuất.",
    "err_date":"Lỗi ngày","err_date2":"Ngày đến phải >= ngày từ.",
    "warn_noday":"Không có ngày giao dịch (T2–T6) trong khoảng đã chọn.",
    "done_ask":"Xuất thành công!\n\nMở file Excel ngay bây giờ?",
    "done_title":"Hoàn thành","no_data":"Không tìm thấy dữ liệu.",
    "cal_today":"Hôm nay","cal_ok":"✓ Chọn",
    "cal_months":["Tháng 1","Tháng 2","Tháng 3","Tháng 4","Tháng 5","Tháng 6",
                  "Tháng 7","Tháng 8","Tháng 9","Tháng 10","Tháng 11","Tháng 12"],
    "cal_wdays":["T2","T3","T4","T5","T6","T7","CN"],
    "lang_btn":"🌐 English","missing_title":"Thiếu thông tin",
    "log_records":"bản ghi","log_nodata":"không có dữ liệu",
    "log_exporting":"Đang xuất Excel...","log_done":"Hoàn thành!",
    "log_skipped":"Bỏ qua","log_skip2":"ngày không có dữ liệu",
    "status_done":"✅ Xong","status_exporting":"Đang xuất Excel...",
    "status_loading":"Đang tải...","err_prefix":"Lỗi",
    "log_period":"Kỳ","log_days":"Ngày","log_types":"Loại GD","log_workers":"Workers",
},
"en": {
    "app_title":"HNX Bond Scraper",
    "mode_single":"Single date","mode_range":"Date range","mode_month":"Full month",
    "lbl_date":"Settlement:","lbl_from":"From:","lbl_to":"To:",
    "lbl_month":"Month:","lbl_year":"Year:",
    "tx_title":"Transaction Types",
    "sel_all":"Select all","clr_all":"Deselect",
    "out_lbl":"Excel file:","browse":"Browse...",
    "workers_lbl":"Parallel:","workers_hint":"days/run",
    "run_btn":"▶  Start Download","running_btn":"⏳  Downloading...",
    "ready":"Ready",
    "warn_type":"Select at least one transaction type.",
    "warn_out":"Please choose an output file.",
    "err_date":"Date error","err_date2":"End date must be >= start date.",
    "warn_noday":"No trading days (Mon–Fri) in selected range.",
    "done_ask":"Export successful!\n\nOpen Excel file now?",
    "done_title":"Done","no_data":"No data found.",
    "cal_today":"Today","cal_ok":"✓ Select",
    "cal_months":["January","February","March","April","May","June",
                  "July","August","September","October","November","December"],
    "cal_wdays":["Mo","Tu","We","Th","Fr","Sa","Su"],
    "lang_btn":"🌐 Tiếng Việt","missing_title":"Missing info",
    "log_records":"records","log_nodata":"no data",
    "log_exporting":"Exporting Excel...","log_done":"Done!",
    "log_skipped":"Skipped","log_skip2":"days with no data",
    "status_done":"✅ Complete","status_exporting":"Exporting Excel...",
    "status_loading":"Loading...","err_prefix":"Error",
    "log_period":"Period","log_days":"Days","log_types":"Types","log_workers":"Workers",
}}

_LANG = ["vi"]
def T(k): return LANG[_LANG[0]].get(k, k)


# ════════════════════════════════════════════════════════════════════════════════
#  Calendar Popup
# ════════════════════════════════════════════════════════════════════════════════

class CalendarPopup(tk.Toplevel):
    def __init__(self, anchor_widget, initial: date, callback):
        super().__init__(anchor_widget)
        self.callback = callback
        self.overrideredirect(True)
        self.configure(bg="#e8ecf5")
        self.resizable(False, False)
        self._yr, self._mo, self._sel = initial.year, initial.month, initial
        self.C = dict(bg="#ffffff", hdr="#1F4E79", hdr_fg="#ffffff",
                      wday="#6b7899", day_fg="#1a1a2e", hover="#dde8f8",
                      sel_bg="#2dce7c", sel_fg="#ffffff",
                      today_bg="#EBF3FB", today_fg="#1F4E79",
                      wknd="#c0392b", sep="#dde2f0", btn="#2E75B6")
        self._build()
        self._position(anchor_widget)
        self.grab_set(); self.focus_force()
        self.bind("<Escape>", lambda e: self.destroy())

    def _position(self, w):
        self.update_idletasks()
        x = w.winfo_rootx()
        y = w.winfo_rooty() + w.winfo_height() + 3
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        rw, rh = self.winfo_reqwidth(), self.winfo_reqheight()
        if x + rw > sw: x = max(0, sw - rw - 4)
        if y + rh > sh: y = w.winfo_rooty() - rh - 3
        self.geometry(f"+{x}+{y}")

    def _build(self):
        C = self.C
        wrap = tk.Frame(self, bg=C["sep"], padx=1, pady=1)
        wrap.pack(fill="both", expand=True)
        box = tk.Frame(wrap, bg=C["bg"])
        box.pack(fill="both", expand=True)

        # Header
        h = tk.Frame(box, bg=C["hdr"], pady=9)
        h.pack(fill="x")
        tk.Button(h, text="◀", bg=C["hdr"], fg=C["hdr_fg"], relief="flat", bd=0,
                  font=("Segoe UI",10,"bold"), activebackground=C["btn"],
                  command=self._prev, cursor="hand2").pack(side="left", padx=10)
        self._hl = tk.Label(h, bg=C["hdr"], fg=C["hdr_fg"], font=("Segoe UI",11,"bold"))
        self._hl.pack(side="left", expand=True)
        tk.Button(h, text="▶", bg=C["hdr"], fg=C["hdr_fg"], relief="flat", bd=0,
                  font=("Segoe UI",10,"bold"), activebackground=C["btn"],
                  command=self._next, cursor="hand2").pack(side="right", padx=10)

        # Weekday strip
        wf = tk.Frame(box, bg="#f0f4ff", pady=5)
        wf.pack(fill="x")
        for i, d in enumerate(T("cal_wdays")):
            tk.Label(wf, text=d, font=("Segoe UI",8,"bold"), bg="#f0f4ff",
                     fg=C["wknd"] if i>=5 else C["wday"],
                     width=3, anchor="center").grid(row=0, column=i, padx=3)

        tk.Frame(box, bg=C["sep"], height=1).pack(fill="x")

        # Grid
        gf = tk.Frame(box, bg=C["bg"], pady=4)
        gf.pack(fill="x", padx=8)
        self._cells = []
        for r in range(6):
            row = []
            for c in range(7):
                lbl = tk.Label(gf, font=("Segoe UI",10), width=3,
                               bg=C["bg"], cursor="hand2", anchor="center", pady=3)
                lbl.grid(row=r, column=c, padx=2, pady=1)
                lbl.bind("<Enter>",    lambda e,b=lbl: self._hov(b, True))
                lbl.bind("<Leave>",    lambda e,b=lbl: self._hov(b, False))
                lbl.bind("<Button-1>", lambda e,b=lbl: self._clk(b))
                row.append(lbl)
            self._cells.append(row)

        tk.Frame(box, bg=C["sep"], height=1).pack(fill="x", padx=8)

        bf = tk.Frame(box, bg=C["bg"], pady=7)
        bf.pack(fill="x", padx=8)
        tk.Button(bf, text=T("cal_today"), font=("Segoe UI",9), bg="#f0f4ff",
                  fg=C["btn"], relief="flat", bd=1, padx=8, pady=4,
                  cursor="hand2", command=self._today).pack(side="left")
        tk.Button(bf, text=T("cal_ok"), font=("Segoe UI",9,"bold"), bg=C["sel_bg"],
                  fg=C["sel_fg"], relief="flat", bd=0, padx=12, pady=4,
                  cursor="hand2", command=self._confirm).pack(side="right")
        self._render()

    def _render(self):
        C = self.C
        self._hl.config(text=f"{T('cal_months')[self._mo-1]}  {self._yr}")
        cal = calendar.monthcalendar(self._yr, self._mo)
        today = date.today()
        for r, row in enumerate(self._cells):
            week = cal[r] if r < len(cal) else [0]*7
            for c, lbl in enumerate(row):
                day = week[c]; lbl._day = day
                if day == 0:
                    lbl.config(text="", bg=C["bg"], relief="flat", highlightthickness=0)
                    continue
                d = date(self._yr, self._mo, day)
                if d == self._sel:       bg,fg,fw = C["sel_bg"],C["sel_fg"],"bold"
                elif d == today:         bg,fg,fw = C["today_bg"],C["today_fg"],"bold"
                else:
                    bg = C["bg"]
                    fg = C["wknd"] if c>=5 else C["day_fg"]
                    fw = "normal"
                lbl.config(text=str(day), bg=bg, fg=fg,
                           font=("Segoe UI",10,fw), relief="flat", highlightthickness=0)

    def _hov(self, lbl, on):
        if not getattr(lbl,"_day",0): return
        if date(self._yr,self._mo,lbl._day)==self._sel: return
        lbl.config(bg=self.C["hover"] if on else self.C["bg"])

    def _clk(self, lbl):
        if not getattr(lbl,"_day",0): return
        self._sel = date(self._yr,self._mo,lbl._day); self._render()

    def _prev(self):
        self._mo -= 1
        if self._mo < 1: self._mo,self._yr = 12,self._yr-1
        self._render()

    def _next(self):
        self._mo += 1
        if self._mo > 12: self._mo,self._yr = 1,self._yr+1
        self._render()

    def _today(self):
        t = date.today(); self._yr,self._mo,self._sel = t.year,t.month,t; self._render()

    def _confirm(self):
        self.callback(self._sel); self.destroy()


# ════════════════════════════════════════════════════════════════════════════════
#  DatePicker widget
# ════════════════════════════════════════════════════════════════════════════════

class DatePicker(tk.Frame):
    def __init__(self, master, initial: date = None, bg="#ffffff"):
        super().__init__(master, bg=bg)
        self._date = initial or date.today()
        self._var  = tk.StringVar(value=self._date.strftime("%d/%m/%Y"))
        self._entry = tk.Entry(self, textvariable=self._var, width=10,
                               font=("Segoe UI",10), relief="solid", bd=1,
                               highlightthickness=2, highlightcolor="#2E75B6",
                               highlightbackground="#dde2f0",
                               bg="#ffffff", fg="#1a1a2e")
        self._entry.pack(side="left")
        self._entry.bind("<FocusOut>", self._parse)
        self._entry.bind("<Return>",   self._parse)
        self._btn = tk.Button(self, text="📅", font=("Segoe UI",9),
                              relief="solid", bd=1, bg="#f0f4ff", fg="#2E75B6",
                              activebackground="#EBF3FB", highlightthickness=0,
                              cursor="hand2", padx=4, pady=2,
                              command=lambda: CalendarPopup(self._btn, self._date, self._pick))
        self._btn.pack(side="left", padx=(2,0))

    def _pick(self, d: date):
        self._date = d
        self._var.set(d.strftime("%d/%m/%Y"))

    def _parse(self, _=None):
        raw = self._var.get().strip()
        for fmt in ("%d/%m/%Y","%d-%m-%Y","%Y-%m-%d"):
            try:
                self._date = datetime.strptime(raw,fmt).date()
                self._var.set(self._date.strftime("%d/%m/%Y"))
                return
            except ValueError: pass
        self._var.set(self._date.strftime("%d/%m/%Y"))

    def get_date(self) -> date:
        self._parse(); return self._date

    def get(self) -> str:
        return self._date.strftime("%d/%m/%Y")


# ════════════════════════════════════════════════════════════════════════════════
#  Application — horizontal wide layout
# ════════════════════════════════════════════════════════════════════════════════

class HNXApp:
    # ── Palette ──
    BG      = "#f0f2f8"
    SURFACE = "#ffffff"
    BORDER  = "#dde2f0"
    ACCENT  = "#1F4E79"
    BLUE    = "#2E75B6"
    GREEN   = "#2dce7c"
    GREEN2  = "#1aab60"
    TEXT    = "#1a1a2e"
    TEXT2   = "#5a6480"
    TEXT3   = "#9aa0b5"
    HOVER   = "#EBF3FB"

    def __init__(self, root):
        self.root = root
        self.root.configure(bg=self.BG)
        self.root.resizable(True, False)
        self._running = False
        self._dyn = {}   # label widgets keyed by T() key for relabelling
        self._build()

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _sep(self, parent, orient="h"):
        if orient == "v":
            tk.Frame(parent, bg=self.BORDER, width=1).pack(
                side="left", fill="y", padx=8, pady=6)
        else:
            tk.Frame(parent, bg=self.BORDER, height=1).pack(fill="x", pady=4)

    def _panel(self, parent, title_key, side="left", fill="both", expand=False, padx=0, pady=0):
        """Titled white panel card."""
        outer = tk.Frame(parent, bg=self.SURFACE,
                         highlightthickness=1, highlightbackground=self.BORDER)
        outer.pack(side=side, fill=fill, expand=expand, padx=padx, pady=pady)
        tl = tk.Label(outer, text=T(title_key), bg=self.ACCENT, fg="#ffffff",
                      font=("Segoe UI",9,"bold"), anchor="w", padx=10, pady=5)
        tl.pack(fill="x")
        self._dyn[title_key] = tl
        body = tk.Frame(outer, bg=self.SURFACE)
        body.pack(fill="both", expand=True, padx=10, pady=8)
        return body

    def _lbl(self, parent, key, row=None, col=None, side=None, **kw):
        lbl = tk.Label(parent, text=T(key), bg=self.SURFACE, fg=self.TEXT2,
                       font=("Segoe UI",9), **kw)
        if row is not None: lbl.grid(row=row, column=col, sticky="w", padx=(0,6), pady=2)
        elif side:          lbl.pack(side=side, padx=(0,6))
        self._dyn[key] = lbl
        return lbl

    def _btn(self, parent, key, cmd, bg=None, fg="#ffffff", **kw):
        bg = bg or self.BLUE
        b = tk.Button(parent, text=T(key), bg=bg, fg=fg, font=("Segoe UI",9),
                      relief="flat", bd=0, cursor="hand2",
                      activebackground=self.ACCENT, activeforeground="#ffffff",
                      padx=10, pady=5, command=cmd, **kw)
        self._dyn[key] = b
        return b

    # ── Build ─────────────────────────────────────────────────────────────────
    def _build(self):
        root = self.root
        root.columnconfigure(0, weight=1)

        # ── TOP BAR ──────────────────────────────────────────────────────────
        top = tk.Frame(root, bg=self.ACCENT)
        top.grid(row=0, column=0, sticky="ew")
        top.columnconfigure(1, weight=1)

        tk.Label(top, text="HNX", bg=self.ACCENT, fg=self.GREEN,
                 font=("Segoe UI",14,"bold")).pack(side="left", padx=(14,4), pady=10)
        tk.Label(top, text="Bond Scraper", bg=self.ACCENT, fg="#ffffff",
                 font=("Segoe UI",14,"bold")).pack(side="left", pady=10)
        tk.Label(top, text="  ·  hnx.vn", bg=self.ACCENT, fg="#7fb3d3",
                 font=("Segoe UI",9)).pack(side="left")

        self._lang_btn = tk.Button(top, text=T("lang_btn"),
                                    font=("Segoe UI",8), bg="#26589a", fg="#ffffff",
                                    relief="flat", padx=8, pady=4, cursor="hand2",
                                    activebackground="#1a3d6e",
                                    command=self._toggle_lang)
        self._lang_btn.pack(side="right", padx=12, pady=8)

        # ── ROW 1: Date mode + Date inputs ───────────────────────────────────
        row1 = tk.Frame(root, bg=self.BG)
        row1.grid(row=1, column=0, sticky="ew", padx=10, pady=(10,0))
        row1.columnconfigure(1, weight=1)

        # Mode selector (left narrow panel)
        mode_body = self._panel(row1, "mode_single", side="left", fill="y", padx=(0,6))
        # Reuse as "Date Mode" title
        self._dyn["mode_single"].config(text="  📅  " + ("Chế độ" if _LANG[0]=="vi" else "Date Mode"))

        self.date_mode = tk.StringVar(value="single")
        self._rb = {}
        for val, key in [("single","mode_single"),("range","mode_range"),("month","mode_month")]:
            rb = tk.Radiobutton(mode_body, text=T(key), variable=self.date_mode, value=val,
                                command=self._mode_changed,
                                bg=self.SURFACE, fg=self.TEXT, activebackground=self.SURFACE,
                                selectcolor=self.BLUE, font=("Segoe UI",10), cursor="hand2",
                                anchor="w")
            rb.pack(fill="x", pady=2)
            self._rb[key] = rb

        # Date input (right expanding panel)
        date_outer = tk.Frame(row1, bg=self.SURFACE,
                              highlightthickness=1, highlightbackground=self.BORDER)
        date_outer.pack(side="left", fill="both", expand=True)
        date_tl = tk.Label(date_outer, text="  🗓  " + T("lbl_date").rstrip(":"),
                            bg=self.BLUE, fg="#ffffff",
                            font=("Segoe UI",9,"bold"), anchor="w", padx=10, pady=5)
        date_tl.pack(fill="x")
        self._date_tl = date_tl
        self._date_body = tk.Frame(date_outer, bg=self.SURFACE)
        self._date_body.pack(fill="both", expand=True, padx=10, pady=10)

        # Single
        self._fr_single = tk.Frame(self._date_body, bg=self.SURFACE)
        tk.Label(self._fr_single, text=T("lbl_date"), bg=self.SURFACE, fg=self.TEXT2,
                 font=("Segoe UI",9)).pack(side="left", padx=(0,8))
        self.dp_single = DatePicker(self._fr_single, date.today(), bg=self.SURFACE)
        self.dp_single.pack(side="left")

        # Range
        self._fr_range = tk.Frame(self._date_body, bg=self.SURFACE)
        self._lbl_from_w = tk.Label(self._fr_range, text=T("lbl_from"), bg=self.SURFACE,
                                     fg=self.TEXT2, font=("Segoe UI",9))
        self._lbl_from_w.pack(side="left", padx=(0,6))
        self.dp_from = DatePicker(self._fr_range, date.today().replace(day=1), bg=self.SURFACE)
        self.dp_from.pack(side="left")
        tk.Label(self._fr_range, text="  →  ", bg=self.SURFACE, fg=self.TEXT3,
                 font=("Segoe UI",13)).pack(side="left")
        self._lbl_to_w = tk.Label(self._fr_range, text=T("lbl_to"), bg=self.SURFACE,
                                   fg=self.TEXT2, font=("Segoe UI",9))
        self._lbl_to_w.pack(side="left", padx=(0,6))
        self.dp_to = DatePicker(self._fr_range, date.today(), bg=self.SURFACE)
        self.dp_to.pack(side="left")

        # Month
        self._fr_month = tk.Frame(self._date_body, bg=self.SURFACE)
        self._lbl_month_w = tk.Label(self._fr_month, text=T("lbl_month"), bg=self.SURFACE,
                                      fg=self.TEXT2, font=("Segoe UI",9))
        self._lbl_month_w.pack(side="left", padx=(0,6))
        self.month_var = tk.StringVar(value=f"{date.today().month:02d}")
        ttk.Combobox(self._fr_month, textvariable=self.month_var,
                     values=[f"{i:02d}" for i in range(1,13)],
                     width=4, state="readonly", font=("Segoe UI",10)).pack(side="left")
        tk.Label(self._fr_month, text=" / ", bg=self.SURFACE, fg=self.TEXT2,
                 font=("Segoe UI",10)).pack(side="left")
        self._lbl_year_w = tk.Label(self._fr_month, text=T("lbl_year"), bg=self.SURFACE,
                                     fg=self.TEXT2, font=("Segoe UI",9))
        self._lbl_year_w.pack(side="left", padx=(0,4))
        self.year_var = tk.StringVar(value=str(date.today().year))
        ttk.Spinbox(self._fr_month, from_=2020, to=2035, textvariable=self.year_var,
                    width=6, font=("Segoe UI",10)).pack(side="left")

        self._mode_changed()

        # ── ROW 2: Transaction types + Output + Settings ──────────────────────
        row2 = tk.Frame(root, bg=self.BG)
        row2.grid(row=2, column=0, sticky="ew", padx=10, pady=(8,0))
        row2.columnconfigure(0, weight=1)

        # Left: transaction types
        tx_outer = tk.Frame(row2, bg=self.SURFACE,
                            highlightthickness=1, highlightbackground=self.BORDER)
        tx_outer.pack(side="left", fill="both", expand=True, padx=(0,6))
        tk.Label(tx_outer, text=f"  ✅  {T('tx_title')}", bg=self.ACCENT, fg="#ffffff",
                 font=("Segoe UI",9,"bold"), anchor="w",
                 padx=10, pady=5).pack(fill="x")
        self._tx_title_lbl = tx_outer.winfo_children()[-1]
        tx_body = tk.Frame(tx_outer, bg=self.SURFACE)
        tx_body.pack(fill="both", expand=True, padx=10, pady=8)

        self.tx_vars = {}
        self._tx_cbs = {}
        for i, (key, label) in enumerate(TRANSACTION_TYPES.items()):
            var = tk.BooleanVar(value=True)
            self.tx_vars[key] = var
            bg  = self.HOVER if i%2==0 else self.SURFACE
            fr  = tk.Frame(tx_body, bg=bg,
                           highlightthickness=1, highlightbackground=self.BORDER)
            fr.grid(row=i//2, column=i%2, sticky="ew",
                    padx=(0 if i%2==0 else 4, 0), pady=2)
            cb = tk.Checkbutton(fr, text=label, variable=var,
                                bg=bg, fg=self.TEXT, activebackground=bg,
                                selectcolor=self.BLUE, font=("Segoe UI",10),
                                cursor="hand2", anchor="w", padx=8, pady=5)
            cb.pack(fill="x")
            self._tx_cbs[key] = cb
        tx_body.columnconfigure(0, weight=1)
        tx_body.columnconfigure(1, weight=1)

        tx_btns = tk.Frame(tx_body, bg=self.SURFACE)
        tx_btns.grid(row=2, column=0, columnspan=2, pady=(6,0), sticky="w")
        self._btn_sel = self._btn(tx_btns, "sel_all",
                                   lambda: [v.set(True) for v in self.tx_vars.values()])
        self._btn_sel.pack(side="left", padx=(0,6))
        self._btn_clr = self._btn(tx_btns, "clr_all",
                                   lambda: [v.set(False) for v in self.tx_vars.values()],
                                   bg="#e8ecf5", fg=self.TEXT2)
        self._btn_clr.pack(side="left")

        # Right: output + settings
        right = tk.Frame(row2, bg=self.BG)
        right.pack(side="left", fill="y")

        out_outer = tk.Frame(right, bg=self.SURFACE,
                              highlightthickness=1, highlightbackground=self.BORDER)
        out_outer.pack(fill="x", pady=(0,6))
        tk.Label(out_outer, text=f"  💾  {T('out_lbl').rstrip(':')}", bg=self.ACCENT,
                 fg="#ffffff", font=("Segoe UI",9,"bold"),
                 anchor="w", padx=10, pady=5).pack(fill="x")
        self._out_tl = out_outer.winfo_children()[-1]
        out_body = tk.Frame(out_outer, bg=self.SURFACE)
        out_body.pack(fill="x", padx=10, pady=8)
        out_body.columnconfigure(0, weight=1)

        self.output_var = tk.StringVar(value=os.path.join(
            os.path.expanduser("~"), "Desktop",
            f"HNX_TraiPhieu_{date.today():%d%m%Y}.xlsx"))

        tk.Entry(out_body, textvariable=self.output_var, width=28,
                 font=("Segoe UI",9), relief="solid", bd=1,
                 highlightthickness=1, highlightbackground=self.BORDER,
                 bg="#fafbff", fg=self.TEXT2).grid(row=0, column=0, sticky="ew", ipady=4)
        self._btn_browse = tk.Button(out_body, text=T("browse"), font=("Segoe UI",9),
                                      bg=self.BLUE, fg="#ffffff", relief="flat",
                                      padx=8, pady=4, cursor="hand2",
                                      activebackground=self.ACCENT,
                                      command=self._browse)
        self._btn_browse.grid(row=0, column=1, padx=(5,0))

        # Workers
        wk_fr = tk.Frame(out_body, bg=self.SURFACE)
        wk_fr.grid(row=1, column=0, columnspan=2, sticky="w", pady=(8,0))
        self._lbl_wk = tk.Label(wk_fr, text=T("workers_lbl"), bg=self.SURFACE,
                                  fg=self.TEXT2, font=("Segoe UI",9))
        self._lbl_wk.pack(side="left")
        self.workers_var = tk.IntVar(value=4)
        ttk.Spinbox(wk_fr, from_=1, to=8, textvariable=self.workers_var,
                    width=3, font=("Segoe UI",10)).pack(side="left", padx=(6,6))
        self._lbl_wk_hint = tk.Label(wk_fr, text=T("workers_hint"), bg=self.SURFACE,
                                      fg=self.TEXT3, font=("Segoe UI",8))
        self._lbl_wk_hint.pack(side="left")

        # ── RUN BUTTON ───────────────────────────────────────────────────────
        self.btn_run = tk.Button(root, text=T("run_btn"),
                                  font=("Segoe UI",12,"bold"),
                                  bg=self.GREEN, fg="#000000",
                                  activebackground=self.GREEN2, activeforeground="#000000",
                                  relief="flat", pady=12, cursor="hand2",
                                  command=self._start)
        self.btn_run.grid(row=3, column=0, sticky="ew", padx=10, pady=8)

        # ── PROGRESS + LOG ───────────────────────────────────────────────────
        bot = tk.Frame(root, bg=self.BG)
        bot.grid(row=4, column=0, sticky="nsew", padx=10, pady=(0,10))
        bot.columnconfigure(0, weight=1)
        root.rowconfigure(4, weight=1)

        # Status + progress bar
        stat_fr = tk.Frame(bot, bg=self.SURFACE,
                           highlightthickness=1, highlightbackground=self.BORDER)
        stat_fr.pack(fill="x", pady=(0,6))
        self.status_var = tk.StringVar(value=T("ready"))
        tk.Label(stat_fr, textvariable=self.status_var, bg=self.SURFACE,
                 fg=self.BLUE, font=("Segoe UI",9,"bold"),
                 anchor="w", padx=10, pady=5).pack(fill="x")
        self._pcanv = tk.Canvas(stat_fr, height=5, bg=self.BORDER,
                                 bd=0, highlightthickness=0)
        self._pcanv.pack(fill="x", padx=0)
        self._pbar = self._pcanv.create_rectangle(0, 0, 0, 5,
                                                    fill=self.GREEN, outline="")
        self._pcanv.bind("<Configure>", self._reprog)
        self._ptotal = 1; self._pval = 0

        # Log
        log_fr = tk.Frame(bot, bg=self.SURFACE,
                          highlightthickness=1, highlightbackground=self.BORDER)
        log_fr.pack(fill="both", expand=True)
        tk.Label(log_fr, text="  📋  Log", bg=self.ACCENT, fg="#ffffff",
                 font=("Segoe UI",9,"bold"), anchor="w",
                 padx=10, pady=5).pack(fill="x")
        inner = tk.Frame(log_fr, bg="#12141e")
        inner.pack(fill="both", expand=True)
        self.log = tk.Text(inner, height=7, font=("Consolas",9),
                           bg="#12141e", fg="#c9d1d9", relief="flat",
                           padx=10, pady=8, state="disabled", wrap="none")
        sb = tk.Scrollbar(inner, command=self.log.yview,
                           bg="#1e2030", troughcolor="#12141e", relief="flat")
        self.log.configure(yscrollcommand=sb.set)
        self.log.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        for t,c in [("ok","#2dce7c"),("err","#ff7b72"),("warn","#e3b341"),
                    ("info","#58a6ff"),("dim","#484f58"),("ts","#484f58")]:
            self.log.tag_configure(t, foreground=c)

        root.title(T("app_title"))
        root.geometry("780x580")
        root.minsize(680, 540)

    # ── Mode ──────────────────────────────────────────────────────────────────
    def _mode_changed(self):
        mode = self.date_mode.get()
        for fr in (self._fr_single, self._fr_range, self._fr_month):
            fr.pack_forget()
        {"single": self._fr_single,
         "range":  self._fr_range,
         "month":  self._fr_month}[mode].pack(anchor="w")

    # ── Language ──────────────────────────────────────────────────────────────
    def _toggle_lang(self):
        _LANG[0] = "en" if _LANG[0] == "vi" else "vi"
        self._relabel()

    def _relabel(self):
        self.root.title(T("app_title"))
        self._lang_btn.config(text=T("lang_btn"))
        # Mode panel title
        mode_title = "  📅  " + ("Chế độ" if _LANG[0]=="vi" else "Date Mode")
        list(self._dyn.get("mode_single",None) and [self._dyn["mode_single"].config(text=mode_title)] or [])
        for key, rb in self._rb.items():
            rb.config(text=T(key))
        self._date_tl.config(text="  🗓  " + (T("lbl_date").rstrip(":")))
        self._lbl_from_w.config(text=T("lbl_from"))
        self._lbl_to_w.config(text=T("lbl_to"))
        self._lbl_month_w.config(text=T("lbl_month"))
        self._lbl_year_w.config(text=T("lbl_year"))
        self._tx_title_lbl.config(text=f"  ✅  {T('tx_title')}")
        self._btn_sel.config(text=T("sel_all"))
        self._btn_clr.config(text=T("clr_all"))
        self._out_tl.config(text=f"  💾  {T('out_lbl').rstrip(':')}")
        self._btn_browse.config(text=T("browse"))
        self._lbl_wk.config(text=T("workers_lbl"))
        self._lbl_wk_hint.config(text=T("workers_hint"))
        if not self._running:
            self.btn_run.config(text=T("run_btn"))
            self.status_var.set(T("ready"))

    # ── Progress ──────────────────────────────────────────────────────────────
    def _reprog(self, _=None):
        w = self._pcanv.winfo_width()
        self._pcanv.coords(self._pbar, 0, 0, int(w*self._pval/max(1,self._ptotal)), 5)

    def _setprog(self, v, total=None):
        if total is not None: self._ptotal = max(1,total)
        self._pval = v
        self.root.after(0, self._reprog)

    # ── Browse ────────────────────────────────────────────────────────────────
    def _browse(self):
        p = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                          filetypes=[("Excel","*.xlsx")],
                                          initialfile=os.path.basename(self.output_var.get()),
                                          initialdir=os.path.dirname(self.output_var.get()))
        if p: self.output_var.set(p)

    # ── Log ───────────────────────────────────────────────────────────────────
    def _log(self, msg, tag=""):
        def _do():
            self.log.configure(state="normal")
            icons = {"ok":"✅ ","err":"❌ ","warn":"⚠ ","info":"ℹ "}
            self.log.insert("end", datetime.now().strftime("%H:%M:%S "),"ts")
            self.log.insert("end", icons.get(tag,"")+msg+"\n", tag or "dim")
            self.log.see("end"); self.log.configure(state="disabled")
        self.root.after(0, _do)

    def _setstatus(self, msg):
        self.root.after(0, lambda: self.status_var.set(msg))

    # ── Start ─────────────────────────────────────────────────────────────────
    def _start(self):
        if self._running: return
        types = [k for k,v in self.tx_vars.items() if v.get()]
        if not types:
            messagebox.showwarning(T("missing_title"), T("warn_type")); return
        out = self.output_var.get().strip()
        if not out:
            messagebox.showwarning(T("missing_title"), T("warn_out")); return
        try:
            mode = self.date_mode.get()
            if mode == "single":
                d = self.dp_single.get_date()
                dates = [d] if d.weekday()<5 else []
                label = d.strftime("%d/%m/%Y")
            elif mode == "range":
                df,dt = self.dp_from.get_date(), self.dp_to.get_date()
                if dt < df:
                    messagebox.showerror(T("err_date"), T("err_date2")); return
                dates = date_range(df, dt)
                label = f"{df:%d/%m/%Y} – {dt:%d/%m/%Y}"
            else:
                m,y = int(self.month_var.get()), int(self.year_var.get())
                dates = dates_for_month(y, m)
                label = f"T{m:02d}/{y}" if _LANG[0]=="vi" else f"{m:02d}/{y}"
        except ValueError as e:
            messagebox.showerror(T("err_date"), str(e)); return
        if not dates:
            messagebox.showwarning(T("missing_title"), T("warn_noday")); return

        self.log.configure(state="normal"); self.log.delete("1.0","end")
        self.log.configure(state="disabled")
        self._setprog(0, len(dates))
        self._running = True
        self.btn_run.configure(state="disabled", text=T("running_btn"), bg="#9e9e9e")
        threading.Thread(target=self._worker,
                         args=(dates,types,out,label), daemon=True).start()

    # ── Worker ────────────────────────────────────────────────────────────────
    def _worker(self, dates, types, output, label):
        import time as _t
        try:
            self._log(f"{T('log_period')}: {label}", "info")
            self._log(f"{T('log_days')}: {len(dates)}  ·  {T('log_types')}: {len(types)}  ·  {T('log_workers')}: {self.workers_var.get()}", "info")
            t0 = _t.time()
            fetched_all, grand, skipped = [], 0, 0
            results = {}
            done = 0

            def run_one(d):
                sess = make_session()
                td   = d.strftime("%d/%m/%Y")
                return td, fetch_day(sess, td, types), d

            with ThreadPoolExecutor(max_workers=self.workers_var.get()) as ex:
                fmap = {ex.submit(run_one, d): d for d in dates}
                for fut in as_completed(fmap):
                    done += 1; self._setprog(done, len(dates))
                    try:
                        td, fdata, d = fut.result()
                        total = sum(len(v) for v in fdata.values())
                        results[d] = (td, fdata, total)
                        tag = "ok" if total else "warn"
                        suf = f"{total} {T('log_records')}" if total else T("log_nodata")
                        self._log(f"[{done}/{len(dates)}] {td} → {suf}", tag)
                        self._setstatus(f"{T('status_loading')} {done}/{len(dates)}")
                    except Exception as e:
                        results[fmap[fut]] = None
                        self._log(f"[{done}/{len(dates)}] {T('err_prefix')}: {e}", "err")

            for d in dates:
                r = results.get(d)
                if not r: skipped += 1
                else:
                    td, fdata, total = r
                    if total == 0: skipped += 1
                    else: fetched_all.append((td,fdata)); grand += total

            if not fetched_all:
                self._log(T("no_data"), "warn"); self._setstatus(f"⚠ {T('no_data')}"); return

            self._log(T("log_exporting"), "info"); self._setstatus(T("status_exporting"))
            export_excel(fetched_all, output, label, types)

            elapsed = _t.time() - t0
            self._log(f"{T('log_done')}  {grand:,} {T('log_records')}  ·  {elapsed:.1f}s", "ok")
            if skipped: self._log(f"{T('log_skipped')} {skipped} {T('log_skip2')}", "warn")
            self._setstatus(f"{T('status_done')} — {grand:,} {T('log_records')} ({elapsed:.1f}s)")
            self._setprog(len(dates), len(dates))
            self.root.after(0, lambda: self._done(output))
        except Exception as e:
            self._log(f"{T('err_prefix')}: {e}", "err")
            self._setstatus(f"❌ {T('err_prefix')}: {e}")
        finally:
            self._running = False
            self.root.after(0, lambda: self.btn_run.configure(
                state="normal", text=T("run_btn"), bg=self.GREEN))

    def _done(self, path):
        if messagebox.askyesno(T("done_title"), T("done_ask")):
            os.startfile(path) if sys.platform=="win32" else os.system(f'open "{path}"')


# ── Entry ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import tkinter as tk
    try:
        import ttkbootstrap as ttk
        root = ttk.Window(themename="litera")
    except Exception:
        import tkinter.ttk as ttk
        root = tk.Tk()
    HNXApp(root)
    root.mainloop()