"""
HNX Bond Trading Data Scraper  (Playwright edition)
=====================================================
Tải kết quả giao dịch trái phiếu trong ngày từ hnx.vn và xuất ra file Excel.
Dùng Playwright để xử lý trang JavaScript động.

Hỗ trợ 4 loại giao dịch:
  - outright       : Giao dịch Outright
  - repo_buy       : Giao dịch Mua Bán Lại
  - repo_sell      : Giao dịch Bán và Mua Lại
  - bond_lending   : Giao dịch Vay Trái Phiếu

Yêu cầu:
    pip install playwright openpyxl pandas beautifulsoup4 lxml
    playwright install chromium

Chạy:
    python hnx_scraper.py                        # Lấy ngày hôm nay, tất cả loại GD
    python hnx_scraper.py --date 03/04/2026      # Ngày cụ thể
    python hnx_scraper.py --type outright        # Chỉ Outright
    python hnx_scraper.py --date 03/04/2026 --type outright --output ket_qua.xlsx
"""

import argparse
import re
import sys
import time
from collections import defaultdict
from datetime import date, datetime

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

# ─────────────────────────── Constants ───────────────────────────────────────

PAGE_URL = "https://hnx.vn/vi-vn/trai-phieu/ket-qua-gd-trong-ngay.html"

TRANSACTION_TYPES = {
    "outright":     {"label": "Giao dịch Outright",       "tab_idx": 0},
    "repo_buy":     {"label": "Giao dịch Mua Bán Lại",    "tab_idx": 1},
    "repo_sell":    {"label": "Giao dịch Bán và Mua Lại", "tab_idx": 2},
    "bond_lending": {"label": "Giao dịch Vay Trái Phiếu", "tab_idx": 3},
}

# ─────────────────────────── Helpers ─────────────────────────────────────────

def parse_table_html(html: str) -> list[dict]:
    """Parse HTML table into list of row dicts."""
    soup = BeautifulSoup("<table>" + html + "</table>", "lxml")
    table = soup.find("table")
    if not table:
        return []

    headers = [th.get_text(strip=True) for th in table.select("thead th, thead td")]
    rows = []
    for tr in table.select("tbody tr"):
        cells = [td.get_text(strip=True).replace("\xa0", "") for td in tr.find_all("td")]
        if not any(cells):
            continue
        if headers and len(cells) >= len(headers):
            rows.append(dict(zip(headers, cells[:len(headers)])))
        elif cells:
            rows.append({f"col_{i}": v for i, v in enumerate(cells)})
    return rows


def try_numeric(val: str):
    """Try to convert string to int or float.
    HNX uses '.' as thousands separator and ',' as decimal separator (Vietnamese format).
    e.g. '3,9767' -> 3.9767   |   '56.829.000.000' -> 56829000000   |   '500.000' -> 500000
    """
    if not isinstance(val, str):
        return val
    v = val.strip()
    if not v:
        return v

    # Count dots and commas to determine format
    dot_count   = v.count(".")
    comma_count = v.count(",")

    if comma_count == 1 and dot_count == 0:
        # e.g. "3,9767" — Vietnamese decimal
        try:
            return float(v.replace(",", "."))
        except ValueError:
            pass
    elif comma_count == 0 and dot_count >= 1:
        # e.g. "500.000" or "56.829.000.000" — thousands separator, integer
        clean = v.replace(".", "")
        try:
            return int(clean)
        except ValueError:
            pass
    elif comma_count == 1 and dot_count >= 1:
        # e.g. "1.700,50" — thousands dot + decimal comma (rare on HNX)
        clean = v.replace(".", "").replace(",", ".")
        try:
            return float(clean)
        except ValueError:
            pass
    elif comma_count == 0 and dot_count == 0:
        try:
            return int(v)
        except ValueError:
            pass

    return v

# ─────────────────────────── Scraping ────────────────────────────────────────

def set_date_and_search(page, trade_date: str) -> bool:
    """
    Set jQuery datepicker on HNX and click TIM KIEM.
    Uses: click input -> select all -> delete -> type date -> Enter -> jQuery fallback -> click button.
    """
    try:
        # Locate date input
        inp = page.locator("input.hasDatepicker").first
        if inp.count() == 0:
            inp = page.locator("input[type='text']").first
        inp.wait_for(timeout=5000)

        # Click to focus, clear existing value, type new date
        inp.click()
        page.wait_for_timeout(200)
        inp.select_text()
        page.keyboard.press("Control+a")
        page.keyboard.press("Delete")
        page.wait_for_timeout(100)
        inp.type(trade_date, delay=80)
        page.wait_for_timeout(200)
        page.keyboard.press("Enter")
        page.wait_for_timeout(400)

        # jQuery datepicker fallback: set via API if available
        page.evaluate("""(d) => {
            const inp = document.querySelector('input.hasDatepicker, input[type="text"]');
            if (!inp) return;
            if (typeof $ !== 'undefined' && $(inp).data('datepicker')) {
                $(inp).datepicker('setDate', d);
            } else {
                inp.value = d;
                inp.dispatchEvent(new Event('change', {bubbles: true}));
            }
        }""", trade_date)
        page.wait_for_timeout(300)

        # Click TIM KIEM search button
        btn = page.locator("button:has-text('TÌM KIẾM'), button:has-text('Tìm kiếm')").first
        btn.click()
        page.wait_for_timeout(3000)
        return True

    except Exception:
        return False

def verify_date_loaded(page, trade_date: str) -> bool:
    """Check that the table shows trade_date in the Ngày BĐGD column."""
    try:
        # Read first data row
        rows = scrape_current_page(page)
        if not rows:
            return False
        sample = rows[0]
        date_key = next(
            (k for k in sample if any(x in k.lower() for x in ["bđgd", "bdgd", "ngày b"])),
            None
        )
        if not date_key:
            return True  # can't verify, assume OK
        return str(sample.get(date_key, "")).strip() == trade_date
    except Exception:
        return False


def click_tab(page, tab_idx: int):
    """Click on a transaction type tab by index."""
    selectors = [
        ".nav-tabs li", ".tab-header li",
        "ul.tabs li", ".bond-tab li",
        ".tab-list li", "[role='tab']",
    ]
    for sel in selectors:
        tabs = page.locator(sel)
        if tabs.count() > tab_idx:
            tabs.nth(tab_idx).click()
            page.wait_for_timeout(1500)
            return
    # Fallback: click by label text
    labels = ["OUTRIGHT", "MUA BÁN LẠI", "BÁN VÀ MUA LẠI", "VAY TRÁI PHIẾU"]
    try:
        page.get_by_text(labels[tab_idx], exact=False).first.click()
        page.wait_for_timeout(1500)
    except Exception:
        pass


def scrape_current_page(page) -> list[dict]:
    """Parse the visible table on current page state."""
    try:
        page.wait_for_selector("table tbody tr", timeout=8000)
        html = page.locator("table").first.inner_html()
        return parse_table_html(html)
    except PWTimeout:
        return []


def click_next_page(page) -> bool:
    """
    Click the next-page control on HNX using JavaScript.
    After switching to 50/page, numbered buttons disappear — only > remains as <li>/<span>.
    """
    try:
        result = page.evaluate("""() => {
            // Try common pagination next selectors
            const selectors = [
                'li.next a', 'li.next',
                '.pagination li:last-child a', '.pagination li:last-child',
                'a[aria-label="Next"]', 'button[aria-label="Next"]',
                '.page-next', '.next-page', '.pager-next'
            ];
            for (const sel of selectors) {
                const el = document.querySelector(sel);
                if (el) {
                    const parent = el.closest('li') || el;
                    if (!parent.classList.contains('disabled') && !parent.classList.contains('active')) {
                        el.click();
                        return 'selector:' + sel;
                    }
                }
            }
            // Fallback: scan ALL elements for text ">" or "»"
            const all = document.querySelectorAll('li, span, a, button, td');
            for (const el of [...all].reverse()) {
                const txt = (el.innerText || '').trim();
                if (txt === '>' || txt === '»' || txt === '›') {
                    const parent = el.closest('li') || el;
                    if (!parent.classList.contains('disabled')) {
                        el.click();
                        return 'text:' + txt;
                    }
                }
            }
            return null;
        }""")
        if result:
            page.wait_for_timeout(2500)
            return True
    except Exception:
        pass
    return False


def wait_for_new_data(page, old_first_stt: str) -> bool:
    """Wait until table refreshes (first row STT changes)."""
    for _ in range(25):
        try:
            page.wait_for_timeout(200)
            rows = scrape_current_page(page)
            if rows:
                if get_row_key(rows[0]) != old_first_stt:
                    return True
        except Exception:
            pass
    return False


def get_next_page_button(page, current_page: int):
    """Return next-page locator if it exists and is enabled."""
    candidates = [
        f"a:has-text('{current_page + 1}')",
        "a[aria-label='Next'], a[aria-label='next']",
        "li.next:not(.disabled) a",
        "a:has-text('»'), a:has-text('>')",
        ".pagination a:last-child",
    ]
    for sel in candidates:
        loc = page.locator(sel).first
        try:
            if loc.count() > 0 and loc.is_visible():
                return loc
        except Exception:
            continue
    return None


def get_total_records(page) -> int:
    """Extract total record count from 'Tổng số X bản ghi' text."""
    try:
        el = page.locator("text=/Tổng số.*bản ghi/").first
        if el.count() > 0:
            m = re.search(r"(\d+)\s*bản ghi", el.inner_text(timeout=3000))
            if m:
                return int(m.group(1))
    except Exception:
        pass
    return 0


def set_page_size_50(page):
    """Change the HNX page-size select dropdown to maximum available (<= 50)."""
    try:
        sel = page.locator("select").first
        if sel.count() > 0:
            options = sel.evaluate("el => Array.from(el.options).map(o => o.value)")
            best = "50" if "50" in options else ("20" if "20" in options else (options[-1] if options else None))
            if best:
                # Use JS to set value AND fire change event (some frameworks need this)
                sel.evaluate(f"""el => {{
                    el.value = '{best}';
                    el.dispatchEvent(new Event('change', {{ bubbles: true }}));
                }}""")
                page.wait_for_timeout(2500)
                print(f"  (Đã đặt hiển thị {best} bản ghi/trang)")
    except Exception:
        pass


def get_row_key(row: dict) -> str | None:
    """Return unique key for a row based on STT column."""
    stt_key = next((k for k in row if k.strip().upper() in ("STT", "TT", "NO", "SỐ TT")), None)
    if stt_key:
        return str(row[stt_key]).strip()
    return str(list(row.values())[0]).strip()


def scrape_tab(page, tx_type: str, trade_date: str) -> list[dict]:
    info = TRANSACTION_TYPES[tx_type]
    print(f"\n📥 Đang tải: {info['label']} ({trade_date}) ...")

    if info["tab_idx"] > 0:
        click_tab(page, info["tab_idx"])

    # Search with retry: keep clicking TÌM KIẾM until table shows correct date
    for attempt in range(3):
        set_date_and_search(page, trade_date)
        set_page_size_50(page)
        if verify_date_loaded(page, trade_date):
            break
        print(f"  (Thử lại lần {attempt+2}: dữ liệu chưa đúng ngày {trade_date}...)")
        page.wait_for_timeout(1000)
    else:
        # After 3 tries, check if there's genuinely no data for this date
        rows = scrape_current_page(page)
        if rows:
            sample = rows[0]
            date_key = next((k for k in sample if any(x in k.lower() for x in ["bđgd","bdgd","ngày b"])), None)
            actual = str(sample.get(date_key,"")).strip() if date_key else "?"
            if actual != trade_date:
                print(f"  ✗ Không thể load dữ liệu ngày {trade_date} (trang hiển thị {actual}) — bỏ qua")
                return []

    total_records = get_total_records(page)
    if total_records > 0:
        print(f"  Tổng số bản ghi cần lấy: {total_records}")

    all_rows  = []
    seen_keys = set()
    page_num  = 1
    MAX_PAGES = 200

    while page_num <= MAX_PAGES:
        rows = scrape_current_page(page)

        if not rows:
            if page_num == 1:
                print(f"  ⚠ Không tìm thấy dữ liệu")
            break

        # Dedup: skip rows whose key we already have
        new_rows = []
        for r in rows:
            key = get_row_key(r)
            if key and key in seen_keys:
                continue
            if key:
                seen_keys.add(key)
            new_rows.append(r)

        if not new_rows:
            print(f"\n  (Trang {page_num} toàn bộ là bản ghi trùng, dừng lại)")
            break

        all_rows.extend(new_rows)
        pct = f"{len(all_rows)}/{total_records}" if total_records else str(len(all_rows))
        print(f"  Trang {page_num}: +{len(new_rows)} mới | Tổng: {pct}", end="\r")

        if total_records > 0 and len(all_rows) >= total_records:
            print(f"\n  ✅ Đã lấy đủ {len(all_rows)}/{total_records} bản ghi")
            break

        # Remember current first row to detect page refresh
        old_first_stt = get_row_key(rows[0]) if rows else None

        # Click next page via JS
        moved = click_next_page(page)
        if not moved:
            print(f"\n  (Không tìm thấy nút sang trang {page_num + 1}, dừng lại)")
            break

        # Wait until table actually shows new data
        if old_first_stt:
            wait_for_new_data(page, old_first_stt)

        page_num += 1

    # ── Filter: only keep rows where Ngày BĐGD == trade_date ─────────────
    # HNX ignores the date param and returns today's data; we verify here.
    if all_rows:
        # Find the BĐGD date column
        sample_keys = list(all_rows[0].keys())
        date_key = next(
            (k for k in sample_keys if any(x in k.lower() for x in
             ["bđgd", "bdgd", "ngày b", "ngày giao d"])),
            None
        )
        if date_key:
            before   = len(all_rows)
            all_rows = [r for r in all_rows if str(r.get(date_key, "")).strip() == trade_date]
            removed  = before - len(all_rows)
            if removed:
                print(f"  ⚠ Lọc ngày {trade_date}: loại bỏ {removed} dòng không khớp")
            if not all_rows:
                print(f"  ✗ Không có dữ liệu khớp ngày {trade_date} "
                      f"(HNX đang hiển thị dữ liệu ngày khác)")
        else:
            print(f"  ⚠ Không tìm thấy cột Ngày BĐGD — không thể xác nhận ngày dữ liệu")

    print(f"  ✓ Hoàn tất: {len(all_rows)} bản ghi                          ")
    return all_rows


def _make_browser_context(pw, headless=True):
    """Create a reusable browser + page."""
    browser = pw.chromium.launch(headless=headless)
    context = browser.new_context(
        ignore_https_errors=True,
        user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        locale="vi-VN",
    )
    page = context.new_page()
    return browser, page


def _load_hnx(page, trade_date: str):
    """Navigate to HNX and set date."""
    try:
        page.goto(PAGE_URL, timeout=30000, wait_until="networkidle")
    except PWTimeout:
        page.goto(PAGE_URL, timeout=30000, wait_until="domcontentloaded")
    page.wait_for_timeout(2000)
    set_date_and_search(page, trade_date)


def scrape_day(page, trade_date: str, types_to_fetch: list) -> dict:
    """Scrape one day using an already-open page. Navigates to correct date."""
    set_date_and_search(page, trade_date)
    results = {}
    for tx_type in types_to_fetch:
        rows = scrape_tab(page, tx_type, trade_date)
        results[tx_type] = rows
        time.sleep(0.3)
    return results


def scrape_all(trade_date: str, types_to_fetch: list, debug: bool = False) -> dict:
    """Single-day convenience wrapper (opens and closes its own browser)."""
    with sync_playwright() as pw:
        browser, page = _make_browser_context(pw, headless=not debug)
        print(f"🌐 Đang mở trang HNX ...")
        _load_hnx(page, trade_date)

        if debug:
            _debug_pagination(page, trade_date)
            browser.close()
            return {}

        results = scrape_day(page, trade_date, types_to_fetch)
        browser.close()
    return results


def scrape_multi(trading_dates: list, types_to_fetch: list) -> list[tuple]:
    """
    Scrape multiple days using a SINGLE browser session.
    Returns list of (trade_date_str, fetched_dict) for days with data.
    """
    all_fetched  = []
    skipped_days = 0
    grand_total  = 0
    n            = len(trading_dates)

    with sync_playwright() as pw:
        browser, page = _make_browser_context(pw, headless=True)
        print(f"🌐 Mở browser một lần cho toàn bộ {n} ngày ...")

        for i, d in enumerate(trading_dates):
            trade_date = d.strftime("%d/%m/%Y")
            print(f"\n[{i+1}/{n}] {trade_date}")

            try:
                # Reload page for each day to reset state cleanly
                _load_hnx(page, trade_date)
                fetched   = scrape_day(page, trade_date, types_to_fetch)
                day_total = sum(len(v) for v in fetched.values())

                if day_total == 0:
                    print(f"  (Không có dữ liệu — bỏ qua)")
                    skipped_days += 1
                else:
                    all_fetched.append((trade_date, fetched))
                    grand_total += day_total
                    print(f"  → {day_total} bản ghi hợp lệ")

            except Exception as e:
                print(f"  ✗ Lỗi ngày {trade_date}: {e} — bỏ qua")
                skipped_days += 1

            if i < n - 1:
                time.sleep(0.5)

        browser.close()

    print(f"\n  Browser đã đóng.")
    return all_fetched, skipped_days, grand_total


def _debug_pagination(page, trade_date: str):
    """Print pagination HTML and screenshot to help diagnose issues."""
    import os
    print("\n🔍 DEBUG MODE - Phân tích cấu trúc trang...")

    # Screenshot
    shot_path = f"hnx_debug_{trade_date.replace('/','')}.png"
    page.screenshot(path=shot_path, full_page=True)
    print(f"  📸 Screenshot: {shot_path}")

    # Print full pagination HTML
    from bs4 import BeautifulSoup
    html = page.content()
    soup = BeautifulSoup(html, "lxml")

    print("\n--- PAGINATION HTML ---")
    for tag in soup.select("nav, .pagination, ul.pager, [class*='page'], [class*='paging']"):
        print(tag.prettify()[:800])
        print("---")

    print("\n--- TOTAL RECORDS TEXT ---")
    for tag in soup.find_all(string=re.compile(r"bản ghi|Tổng số|total", re.I)):
        print(repr(str(tag).strip()))

    print("\n--- ALL LINKS IN PAGE FOOTER ---")
    for a in soup.select("a[href], button"):
        txt = a.get_text(strip=True)
        if txt and any(c.isdigit() for c in txt) or txt in ("»", ">", ">>", "Next", "Tiếp"):
            print(f"  [{a.name}] text={repr(txt)} class={a.get('class')} href={a.get('href','')[:80]}")

    print("\n--- TABLE ROW COUNT ---")
    rows = page.locator("table tbody tr")
    print(f"  Số dòng trong bảng: {rows.count()}")

    print("\n--- PAGE SIZE CONTROLS ---")
    for tag in soup.select("select, input[type='number']"):
        print(f"  [{tag.name}] name={tag.get('name')} id={tag.get('id')} class={tag.get('class')}")
        if tag.name == "select":
            for opt in tag.find_all("option"):
                print(f"    option: value={opt.get('value')} text={opt.get_text(strip=True)}")


# ─────────────────────────── Excel Export ────────────────────────────────────

NAVY    = "1F4E79"
BLUE    = "2E75B6"
LT_BLUE = "EBF3FB"

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(cell, text, bg=BLUE, fc="FFFFFF", bold=True, size=10):
    cell.value = text
    cell.font  = Font(name="Arial", bold=bold, size=size, color=fc)
    cell.fill  = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border()


# Default fallback headers if no data available (used when sheet is empty)
# Fallback headers only used when a sheet has NO data at all (empty sheet placeholder).
# Real headers are always taken from actual scraped data (records[0].keys()).
# These are best-guess based on HNX website — may differ slightly per transaction type.
FALLBACK_HEADERS = {
    "outright":     ["STT", "Ngày giao kết giao dịch", "Ngày BĐGD", "Kỳ hạn còn lại",
                     "Mã TP", "Tiền tệ", "Giá yết (đồng)", "Lợi suất (%/năm)",
                     "KLGD", "GTGD (đồng)", "Ngày thanh toán"],
    "repo_buy":     ["STT", "Ngày giao kết giao dịch", "Ngày BĐGD", "Kỳ hạn còn lại",
                     "Mã TP", "Tiền tệ", "Giá yết mua lại (đồng)", "Lãi suất (%/năm)",
                     "KLGD", "GTGD (đồng)", "Ngày mua lại", "Ngày thanh toán"],
    "repo_sell":    ["STT", "Ngày giao kết giao dịch", "Ngày BĐGD", "Kỳ hạn còn lại",
                     "Mã TP", "Tiền tệ", "Giá yết bán lại (đồng)", "Lãi suất (%/năm)",
                     "KLGD", "GTGD (đồng)", "Ngày bán lại", "Ngày thanh toán"],
    "bond_lending": ["STT", "Ngày giao kết giao dịch", "Ngày BĐGD", "Kỳ hạn còn lại",
                     "Mã TP", "Tiền tệ", "Giá yết (đồng)", "Phí vay (%/năm)",
                     "KLGD", "GTGD (đồng)", "Ngày hoàn trả", "Ngày thanh toán"],
}

# Cache: stores actual headers discovered from live data, keyed by tx_type
_DISCOVERED_HEADERS: dict = {}


def get_or_create_sheet(wb: Workbook, sheet_name: str, tx_type: str, keys: list) -> tuple:
    """Get existing sheet or create new one with headers. Returns (ws, next_row)."""
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Find next empty row
        next_row = ws.max_row + 1
        return ws, next_row
    else:
        ws = wb.create_sheet(sheet_name)
        # Row 1: column headers
        ws.row_dimensions[1].height = 28
        for ci, k in enumerate(keys, 1):
            _hdr(ws.cell(1, ci), k)
            ws.column_dimensions[get_column_letter(ci)].width = max(14, len(k) + 3)
        ws.freeze_panes = "A2"
        return ws, 2


def append_records(wb: Workbook, sheet_name: str, label: str,
                   records: list[dict], tx_type: str = "outright"):
    """Append records to sheet, creating it with headers if needed."""
    if records:
        keys = list(records[0].keys())
        # Cache the real headers for use in ensure_empty_sheets later
        _DISCOVERED_HEADERS[tx_type] = keys
    else:
        keys = _DISCOVERED_HEADERS.get(tx_type) or FALLBACK_HEADERS.get(tx_type, FALLBACK_HEADERS["outright"])
    ws, start_row = get_or_create_sheet(wb, sheet_name, tx_type, keys)

    for ri, rec in enumerate(records):
        r = start_row + ri
        fill = PatternFill("solid", start_color=LT_BLUE if (r % 2 == 0) else "FFFFFF")
        for ci, k in enumerate(keys, 1):
            cell = ws.cell(r, ci)
            val  = try_numeric(rec.get(k, ""))
            cell.value  = val
            cell.font   = Font(name="Arial", size=10)
            cell.fill   = fill
            cell.border = _border()
            if isinstance(val, (int, float)):
                cell.number_format = "#,##0.####"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")


def ensure_empty_sheets(wb: Workbook):
    """Make sure all 4 transaction type sheets exist, even if empty.
    Uses real discovered headers if available, fallback otherwise."""
    for tx_type, info in TRANSACTION_TYPES.items():
        sheet_name = info["label"][:28]
        if sheet_name not in wb.sheetnames:
            # Prefer headers discovered from actual data; fall back to guesses
            keys = (_DISCOVERED_HEADERS.get(tx_type)
                    or FALLBACK_HEADERS.get(tx_type)
                    or FALLBACK_HEADERS["outright"])
            ws = wb.create_sheet(sheet_name)
            ws.row_dimensions[1].height = 28
            for ci, k in enumerate(keys, 1):
                _hdr(ws.cell(1, ci), k)
                ws.column_dimensions[get_column_letter(ci)].width = max(14, len(k) + 3)
            ws.freeze_panes = "A2"


def write_cover(wb: Workbook, date_range_label: str, summary: dict):
    """Write/update cover sheet with run summary."""
    if "Tổng quan" in wb.sheetnames:
        del wb["Tổng quan"]
    ws = wb.create_sheet("Tổng quan", 0)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14

    ws.merge_cells("A1:D1")
    _hdr(ws["A1"], "KẾT QUẢ GIAO DỊCH TRÁI PHIẾU – HNX", bg=NAVY, size=13)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:D2")
    _hdr(ws["A2"],
         f"Kỳ: {date_range_label}  |  Tạo lúc: {datetime.now():%d/%m/%Y %H:%M}  |  Nguồn: hnx.vn",
         bg=BLUE, size=9, bold=False)
    ws.row_dimensions[2].height = 16

    ws.append([])
    for ci, h in enumerate(["Loại giao dịch", "Sheet", "Số ngày có GD", "Tổng bản ghi"], 1):
        _hdr(ws.cell(4, ci), h)
    ws.row_dimensions[4].height = 22

    for i, (tx_type, info) in enumerate(TRANSACTION_TYPES.items()):
        r = 5 + i
        s = summary.get(tx_type, {"days": 0, "records": 0})
        fill = PatternFill("solid", start_color=LT_BLUE if i % 2 == 0 else "FFFFFF")
        for ci, (val, align) in enumerate([
            (info["label"],       "left"),
            (info["label"][:28],  "center"),
            (s["days"],           "right"),
            (s["records"],        "right"),
        ], 1):
            cell = ws.cell(r, ci)
            cell.value = val
            cell.font  = Font(name="Arial", size=10)
            cell.fill  = fill
            cell.border = _border()
            cell.alignment = Alignment(horizontal=align, vertical="center")
            if ci in (3, 4):
                cell.number_format = "#,##0"


def export_excel_multiday(all_fetched: list[tuple], output_path: str,
                          date_range_label: str):
    """
    Build Excel from multiple days of data.
    all_fetched: list of (trade_date_str, fetched_dict)
    Each tx_type gets one sheet; all dates appended row by row.
    """
    wb = Workbook()
    # Remove default sheet; sheets created on demand
    wb.remove(wb.active)

    summary = {tx: {"days": 0, "records": 0} for tx in TRANSACTION_TYPES}

    for trade_date, fetched in all_fetched:
        for tx_type, info in TRANSACTION_TYPES.items():
            records = fetched.get(tx_type, [])
            sheet_name = info["label"][:28]
            if records:
                append_records(wb, sheet_name, info["label"], records, tx_type=tx_type)
                summary[tx_type]["days"]    += 1
                summary[tx_type]["records"] += len(records)

    ensure_empty_sheets(wb)
    write_cover(wb, date_range_label, summary)

    # Reorder: Tổng quan first, then transaction sheets
    sheet_order = ["Tổng quan"] + [info["label"][:28] for info in TRANSACTION_TYPES.values()]
    for i, name in enumerate(sheet_order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=wb.sheetnames.index(name) - i)

    wb.save(output_path)
    print(f"\n✅ Đã lưu: {output_path}")


# ─────────────────────────── Date helpers ────────────────────────────────────

def parse_date(s: str) -> date:
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Không nhận ra định dạng ngày: {s!r}. Dùng dd/MM/yyyy")


def date_range(start: date, end: date) -> list[date]:
    """Return all Mon-Fri dates from start to end inclusive."""
    from datetime import timedelta
    days = []
    cur = start
    while cur <= end:
        if cur.weekday() < 5:
            days.append(cur)
        cur += timedelta(days=1)
    return days


def dates_for_month(year: int, month: int) -> list[date]:
    import calendar
    _, last_day = calendar.monthrange(year, month)
    start = date(year, month, 1)
    end   = date(year, month, last_day)
    return date_range(start, end)


# ─────────────────────────── CLI ─────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(
        description="Scrape HNX bond trading results → Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )

    # ── Date selection (mutually exclusive) ──
    date_grp = p.add_mutually_exclusive_group()
    date_grp.add_argument("--date", "-d",
                          help="Một ngày cụ thể: dd/MM/yyyy (mặc định: hôm nay)")
    date_grp.add_argument("--from-date", "-f",
                          help="Từ ngày: dd/MM/yyyy (dùng với --to-date)")
    date_grp.add_argument("--month", "-m",
                          help="Cả tháng: MM/yyyy  (ví dụ: 03/2026)")

    p.add_argument("--to-date", "-e",
                   help="Đến ngày: dd/MM/yyyy (dùng với --from-date)")
    p.add_argument("--type", "-t",
                   nargs="+",
                   choices=["all"] + list(TRANSACTION_TYPES.keys()),
                   default=["all"],
                   metavar="TYPE",
                   help="Loại GD, chọn nhiều: -t outright repo_buy | all/outright/repo_buy/repo_sell/bond_lending")
    p.add_argument("--output", "-o", default=None,
                   help="Tên file Excel output")
    p.add_argument("--debug", action="store_true",
                   help="Debug: chụp ảnh + in HTML pagination, không xuất Excel")
    args = p.parse_args()

    # ── Resolve date list ──
    if args.month:
        try:
            m, y = args.month.split("/")
            trading_dates = dates_for_month(int(y), int(m))
            range_label = f"Tháng {args.month}"
            default_out = f"HNX_TraiPhieu_{args.month.replace('/', '')}.xlsx"
        except Exception:
            print("Lỗi: --month phải có dạng MM/yyyy, ví dụ: 03/2026")
            sys.exit(1)

    elif args.from_date:
        if not args.to_date:
            print("Lỗi: --from-date cần đi kèm --to-date")
            sys.exit(1)
        start = parse_date(args.from_date)
        end   = parse_date(args.to_date)
        if end < start:
            print("Lỗi: --to-date phải >= --from-date")
            sys.exit(1)
        trading_dates = date_range(start, end)
        range_label   = f"{start:%d/%m/%Y} – {end:%d/%m/%Y}"
        default_out   = f"HNX_TraiPhieu_{start:%d%m%Y}_{end:%d%m%Y}.xlsx"

    else:
        d = parse_date(args.date) if args.date else date.today()
        trading_dates = [d]
        range_label   = f"{d:%d/%m/%Y}"
        default_out   = f"HNX_TraiPhieu_{d:%d%m%Y}.xlsx"

    output         = args.output or default_out
    selected = args.type  # now a list
    if "all" in selected:
        types_to_fetch = list(TRANSACTION_TYPES.keys())
    else:
        # validate and deduplicate while preserving order
        valid = list(TRANSACTION_TYPES.keys())
        types_to_fetch = [t for t in valid if t in selected]

    print("=" * 65)
    print(f"  HNX Bond Scraper")
    print(f"  Kỳ     : {range_label}")
    print(f"  Số ngày: {len(trading_dates)} ngày giao dịch")
    print(f"  Loại GD: {', '.join(types_to_fetch)}")
    print(f"  Output : {output}")
    print("=" * 65)

    if args.debug:
        # Debug: single day, show pagination info
        d = trading_dates[0]
        scrape_all(d.strftime("%d/%m/%Y"), types_to_fetch, debug=True)
        return

    if len(trading_dates) == 1:
        # Single day: simple path
        d = trading_dates[0]
        trade_date = d.strftime("%d/%m/%Y")
        print(f"\n[1/1] {trade_date}")
        fetched    = scrape_all(trade_date, types_to_fetch)
        day_total  = sum(len(v) for v in fetched.values())
        if day_total == 0:
            print("\n⚠  Không tìm thấy dữ liệu.")
            sys.exit(1)
        all_fetched  = [(trade_date, fetched)]
        skipped_days = 0
        grand_total  = day_total
    else:
        # Multi-day: single browser session
        all_fetched, skipped_days, grand_total = scrape_multi(trading_dates, types_to_fetch)

    if not all_fetched:
        print("\n⚠  Không tìm thấy dữ liệu cho toàn bộ kỳ.")
        sys.exit(1)

    export_excel_multiday(all_fetched, output, range_label)

    print("\n📊 Tóm tắt:")
    print(f"   Kỳ dữ liệu  : {range_label}")
    print(f"   Ngày có GD  : {len(all_fetched)}/{len(trading_dates)}")
    if skipped_days:
        print(f"   Ngày bỏ qua : {skipped_days} (không có dữ liệu / lỗi)")
    print(f"   Tổng bản ghi: {grand_total:,}")
    print(f"   File         : {output}")


if __name__ == "__main__":
    main()